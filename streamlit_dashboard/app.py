import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
import os
import glob
import json
import calendar
import urllib.parse
import time
import streamlit_antd_components as sac
from google_calendar_utils import get_calendar_service, add_event_to_calendar, get_credentials, get_user_info
from google_sheets_utils import GoogleSheetsManager
import app_translations as tr
from app_translations import TRANSLATIONS
from sklearn.preprocessing import LabelEncoder

from sklearn.ensemble import RandomForestRegressor
import ai_utils
from flashcard_data import FLASHCARD_DATA

# Load translations
if "language" not in st.session_state:
    st.session_state.language = "日本語" # Default language

# --- Global CSS Animations ---
st.markdown("""
<style>
/* Fade-in animation for main content */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(10px); }
    to { opacity: 1; transform: translateY(0); }
}
.stApp {
    animation: fadeIn 0.5s ease-out;
}

/* Hover effect for metric cards (if they use st.metric or custom containers) */
div[data-testid="stMetric"] {
    transition: transform 0.2s ease, box-shadow 0.2s ease;
    padding: 10px;
    border-radius: 8px;
}
div[data-testid="stMetric"]:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
    background-color: rgba(255, 255, 255, 0.05); /* Subtle highlight */
}

/* Pulse animation for urgent alerts */
@keyframes pulse-red {
    0% { box-shadow: 0 0 0 0 rgba(239, 68, 68, 0.4); }
    70% { box-shadow: 0 0 0 10px rgba(239, 68, 68, 0); }
    100% { box-shadow: 0 0 0 0 rgba(239, 68, 68, 0); }
}
div[data-testid="stAlert"][class*="danger"] {
    animation: pulse-red 2s infinite;
}

/* Smooth transition for tabs */
div[data-testid="stTabs"] button {
    transition: all 0.3s ease;
}

/* Button Hover Animation (Scale Up) */
div[data-testid="stButton"] button {
    transition: transform 0.1s ease-in-out, box-shadow 0.1s ease;
}
div[data-testid="stButton"] button:hover {
    transform: scale(1.02);
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
}
div[data-testid="stButton"] button:active {
    transform: scale(0.98);
}

/* Chat Message Slide-in Animation */
div[data-testid="stChatMessage"] {
    animation: slideInLeft 0.3s ease-out;
}
@keyframes slideInLeft {
    from { opacity: 0; transform: translateX(-10px); }
    to { opacity: 1; transform: translateX(0); }
}

/* Input Field Focus Transition */
div[data-testid="stTextInput"] input, div[data-testid="stNumberInput"] input {
    transition: border-color 0.3s ease, box-shadow 0.3s ease;
}
div[data-testid="stTextInput"] input:focus, div[data-testid="stNumberInput"] input:focus {
    border-color: #3b82f6;
    box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.2);
}
</style>
""", unsafe_allow_html=True)


# ダッシュボード表示設定の初期化
if "dashboard_widgets_v2" not in st.session_state:
    st.session_state.dashboard_widgets_v2 = ["主要指標", "学習カレンダー", "学習記録", "週間学習プラン"]
elif "週間学習プラン" not in st.session_state.dashboard_widgets_v2:
    st.session_state.dashboard_widgets_v2.append("週間学習プラン")

def t(key):
    return tr.get_text(key, st.session_state.get("language", "日本語"))

def dt(text):
    return tr.get_data_text(text, st.session_state.get("language", "日本語"))

# --- 安全な再実行トリガ（環境差分を吸収） ---
def trigger_rerun():
    """
    Streamlit の再実行を安全に呼び出す。
    """
    try:
        if hasattr(st, "rerun"):
            st.rerun()
        elif hasattr(st, "experimental_rerun"):
            st.experimental_rerun()
        else:
            try:
                qp = dict(st.query_params) if hasattr(st, "query_params") else {}
                qp["_rerun"] = int(time.time())
                st.query_params = qp
            except Exception:
                if hasattr(st, "set_query_params"):
                    st.set_query_params(**qp)
    except Exception:
        return

@st.cache_resource
def train_ai_models(df):
    """
    機械学習モデルの学習（キャッシュ化）
    Random Forestを用いて正答率を予測し、重要変数を抽出する
    """
    # データが少なすぎる場合は学習しない
    if df.empty or len(df) < 5:
        return None, None, None
    
    try:
        # データ前処理
        df_ml = df.copy()
        df_ml["date_obj"] = pd.to_datetime(df_ml["日付"])
        # 基準日からの経過日数
        min_date = df_ml["date_obj"].min()
        df_ml["days_passed"] = (df_ml["date_obj"] - min_date).dt.days
        # 正誤を数値化 (1/0)
        df_ml["is_correct"] = df_ml["正誤"].apply(lambda x: 1 if x == "〇" else 0)
        
        # 欠損値処理
        df_ml = df_ml.fillna(0)
        
        # カテゴリ変数のエンコーディング
        le_subj = LabelEncoder()
        le_unit = LabelEncoder()
        
        # 文字列型に変換してからエンコード
        df_ml["subj_code"] = le_subj.fit_transform(df_ml["科目"].astype(str))
        df_ml["unit_code"] = le_unit.fit_transform(df_ml["単元"].astype(str))
        
        # 特徴量: 経過日数, 科目, 単元, 解答時間, 学習投入時間
        # ※本来はOneHotEncodingすべきだが、決定木ベースなのでLabelEncodingでも許容
        X = df_ml[["days_passed", "subj_code", "unit_code", "解答時間(秒)", "学習投入時間(分)"]]
        y = df_ml["is_correct"]
        
        # モデル学習 (Random Forest Regressor)
        # 0/1の分類ではなく、確率(正答率)として予測したいので回帰モデルを使用
        model = RandomForestRegressor(n_estimators=50, max_depth=5, random_state=42, n_jobs=-1)
        model.fit(X, y)
        
        # 変数重要度
        importances = pd.DataFrame({
            "feature": ["経過日数", "科目", "単元", "解答時間", "学習時間"],
            "importance": model.feature_importances_
        }).sort_values("importance", ascending=False)
        
        return model, importances, (le_subj, le_unit, min_date)
        
    except Exception as e:
        st.error(f"AI学習エラー: {e}")
        return None, None, None

def generate_weekly_study_plan(df, exam_date, target_rate, current_rate):
    """
    週間学習プラン自動生成 (エビングハウス忘却曲線 + 可用時間考慮)
    """
    if exam_date is None:
        return None
    
    # 残り日数計算
    today = datetime.today().date()
    days_left = (exam_date - today).days
    
    if days_left < 0:
        return None
    
    if df.empty:
        return None

    # 1. 復習候補の特定 (エビングハウス忘却曲線: 1, 3, 7, 14, 30日前)
    review_intervals = [1, 3, 7, 14, 30]
    review_candidates = {} # date -> set(units)
    
    # 過去の学習ログから復習すべき単元を特定
    df["date_obj"] = pd.to_datetime(df["日付"]).dt.date
    
    for day in range(min(7, days_left)):
        target_date = today + timedelta(days=day)
        review_units = set()
        
        # この日(target_date)に復習すべき過去の日付を計算
        for interval in review_intervals:
            past_date = target_date - timedelta(days=interval)
            # past_dateに学習した単元を取得
            studied_on_date = df[df["date_obj"] == past_date]["単元"].unique()
            for unit in studied_on_date:
                review_units.add(unit)
        
        review_candidates[target_date] = list(review_units)

    # 2. 弱点単元の抽出
    weak_units = df.groupby("単元").agg({
        "ミス": ["sum", "count"]
    }).reset_index()
    weak_units.columns = ["単元", "ミス数", "試行回数"]
    weak_units["正答率"] = (weak_units["試行回数"] - weak_units["ミス数"]) / weak_units["試行回数"]
    weak_units["優先度"] = (1 - weak_units["正答率"]) * weak_units["試行回数"]
    weak_list = weak_units.sort_values("優先度", ascending=False)["単元"].tolist()
    
    # 3. 週間プラン生成
    weekly_plan = {}
    daily_limit_mins = st.session_state.get("daily_study_time", 60)
    unit_time_mins = 20 # 1単元あたりの想定時間
    
    # 過去7日 + 未来28日 (約1ヶ月)
    start_day = -7
    end_day = min(28, days_left + 1)
    
    for day in range(start_day, end_day):
        date = today + timedelta(days=day)
        date_str = date.strftime("%Y-%m-%d")
        
        todays_units = []
        current_time = 0
        
        if day < 0:
            # 過去: 学習ログから実績を表示
            if not df.empty:
                # date_objは既に作成済みと仮定、なければ作成
                if "date_obj" not in df.columns:
                    df["date_obj"] = pd.to_datetime(df["日付"]).dt.date
                
                day_logs = df[df["date_obj"] == date]
                for _, row in day_logs.iterrows():
                    # 重複排除（同じ単元を複数回やった場合など）
                    if not any(u["name"] == dt(row["単元"]) for u in todays_units):
                        todays_units.append({
                            "name": dt(row["単元"]),
                            "type": t("completed"), # "完了" or similar
                            "subject": row["科目"]
                        })
                        current_time += row.get("学習投入時間(分)", 20) # データがなければ20分仮定
        else:
            # 未来: プラン生成 (既存ロジック)
            
            # A. 復習単元を優先的に追加
            reviews = review_candidates.get(date, [])
            for unit in reviews:
                if current_time + unit_time_mins <= daily_limit_mins:
                    # 科目を特定（dfから）
                    subject = df[df["単元"] == unit]["科目"].iloc[0] if not df[df["単元"] == unit].empty else "復習"
                    todays_units.append({"name": dt(unit), "type": t("plan_review"), "subject": subject})
                    current_time += unit_time_mins
            
            # B. 時間が余っていれば弱点単元を追加
            weak_idx = 0
            while current_time + unit_time_mins <= daily_limit_mins and weak_idx < len(weak_list):
                unit = weak_list[weak_idx]
                # まだリストになければ追加
                if not any(u["name"] == dt(unit) for u in todays_units):
                    subject = df[df["単元"] == unit]["科目"].iloc[0] if not df[df["単元"] == unit].empty else "弱点"
                    todays_units.append({"name": dt(unit), "type": t("plan_weakness"), "subject": subject})
                    current_time += unit_time_mins
                weak_idx += 1
                
            # C. それでも時間が余っていれば、ランダムまたは次の弱点を追加
            while current_time + unit_time_mins <= daily_limit_mins:
                 if weak_idx < len(weak_list):
                    unit = weak_list[weak_idx]
                    if not any(u["name"] == dt(unit) for u in todays_units):
                        subject = df[df["単元"] == unit]["科目"].iloc[0] if not df[df["単元"] == unit].empty else "演習"
                        todays_units.append({"name": dt(unit), "type": t("study"), "subject": subject})
                        current_time += unit_time_mins
                    weak_idx += 1
                 else:
                     break 
            
            # D. 最低限の学習を保証 (時間が埋まってなくても、まだ何もなければ追加)
            if not todays_units and weak_list:
                unit = weak_list[0]
                subject = df[df["単元"] == unit]["科目"].iloc[0] if not df[df["単元"] == unit].empty else "演習"
                todays_units.append({"name": dt(unit), "type": t("plan_weakness"), "subject": subject})
                current_time += unit_time_mins

        if todays_units:
            weekly_plan[date_str] = {
                "units": todays_units,
                "time_minutes": int(current_time)
            }
        else:
             # データがない場合も空エントリを追加して、カレンダー上で日付が表示されるようにする
             weekly_plan[date_str] = {"units": [], "time_minutes": 0}

    return weekly_plan

def generate_ai_advice(current_rate, target_rate, time_excess_rate, streak_days):
    """
    ルールベースAIによる学習アドバイス生成
    """
    advices = []
    
    # 1. 正答率に基づくアドバイス
    if current_rate >= target_rate:
        advices.append(("<i class='bi bi-star-fill' style='color:#fbbf24;'></i>", t("ai_advice_high_accuracy")))
    elif current_rate >= target_rate - 0.1:
        advices.append(("<i class='bi bi-fire' style='color:#f97316;'></i>", t("ai_advice_almost_there")))
    else:
        advices.append(("<i class='bi bi-lightbulb-fill' style='color:#f59e0b;'></i>", t("ai_advice_needs_work")))
        
    # 2. 解答時間に基づくアドバイス
    if time_excess_rate > 0.3:
        advices.append(("<i class='bi bi-stopwatch' style='color:#6b7280;'></i>", t("ai_advice_slow")))
    elif time_excess_rate < 0.1:
        advices.append(("<i class='bi bi-lightning-charge-fill' style='color:#eab308;'></i>", t("ai_advice_fast")))
        
    # 3. 継続日数に基づくアドバイス
    if streak_days >= 3:
        advices.append(("<i class='bi bi-calendar-check-fill' style='color:#ef4444;'></i>", t("ai_advice_streak").format(streak_days)))
    elif streak_days == 0:
        advices.append(("<i class='bi bi-megaphone-fill' style='color:#3b82f6;'></i>", t("ai_advice_no_study")))
        
    # ランダムに1つ、または状況に合わせて結合して返す
    # ここではメインのアドバイス（正答率）とサブアドバイスを組み合わせる
    main_icon, main_text = advices[0]
    
    if len(advices) > 1:
        sub_icon, sub_text = advices[1] if len(advices) > 1 else ("", "")
        return f"**AIコーチ**: {main_icon} {main_text}  \n{sub_icon} {sub_text}"
    else:
        return f"**AIコーチ**: {main_icon} {main_text}"

def generate_calendar_heatmap(df, year, month, exam_date=None, weekly_plan=None):
    """
    学習カレンダーヒートマップを生成（強化版）
    - 単月表示
    - 未来の学習予定表示
    - 試験日ハイライト
    """
    try:
        from datetime import datetime, timedelta
        import calendar as cal
        import pandas as pd # pandas import added for df_copy = pd.DataFrame()
        
        # 日付列を確実にdatetime型に変換
        df_copy = df.copy() if not df.empty else pd.DataFrame()
        if not df_copy.empty:
            df_copy["日付"] = pd.to_datetime(df_copy["日付"], errors='coerce')
            df_copy = df_copy.dropna(subset=["日付"])
            df_copy["日付"] = df_copy["日付"].dt.date
        
        # 日別に集計
        daily_stats_dict = {}
        if not df_copy.empty:
            daily_stats = df_copy.groupby(
                "日付"
            ).agg({
                "問題ID": "count",
                "正誤": lambda x: (x == "〇").mean(),
                "学習投入時間(分)": "sum"
            }).reset_index()
            
            daily_stats.columns = ["日付", "問題数", "正答率", "学習時間"]
            daily_stats_dict = daily_stats.set_index("日付").to_dict('index')
        
        # 週間プランから未来の予定を取得
        future_plan_dict = {}
        if weekly_plan:
            for day_key, units in weekly_plan.items():
                try:
                    # day_keyがすでにdatetime.dateオブジェクトの場合
                    if isinstance(day_key, datetime.date):
                        date_obj = day_key
                    else:
                        # 文字列の場合（"01/23 (Mon)"形式）
                        month_day_str = day_key.split(' ')[0]
                        current_year = datetime.now().year
                        date_obj = datetime.strptime(f"{current_year}/{month_day_str}", "%Y/%m/%d").date()
                    
                    # unitsが辞書でunitsキーを持つ場合
                    if isinstance(units, dict) and "units" in units:
                        future_plan_dict[date_obj] = len(units["units"])
                    # unitsがリストの場合
                    elif isinstance(units, list):
                        future_plan_dict[date_obj] = len(units)
                    # その他の場合は単に存在フラグとして1を設定
                    elif units:
                        future_plan_dict[date_obj] = 1
                except Exception as e:
                    # エラーは無視して次へ
                    pass
        
        # CSSを定義
        css = """
        <style>
        .calendar-single {
            background: white;
            border-radius: 12px;
            padding: 8px 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            max-width: 100%;
            margin: 0 auto;
            font-family: "Source Sans Pro", sans-serif;
        }
        .calendar-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
        }
        .calendar-title {
            font-size: 1.3rem;
            font-weight: 800;
            color: #1f2937;
        }
        .calendar-nav {
            display: flex;
            gap: 8px;
        }
        .calendar-nav-btn {
            background: #f3f4f6;
            border: none;
            border-radius: 6px;
            padding: 8px 12px;
            cursor: pointer;
            font-weight: 600;
            color: #374151;
            transition: all 0.2s;
        }
        .calendar-nav-btn:hover {
            background: #e5e7eb;
        }
        .calendar-table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }
        .calendar-weekday {
            font-size: 0.85rem;
            font-weight: 700;
            color: #6b7280;
            text-align: center;
            padding: 12px 8px;
            border-bottom: 2px solid #e5e7eb;
        }
        .calendar-day {
            aspect-ratio: 1;
            text-align: center;
            vertical-align: middle;
            font-size: 0.9rem;
            cursor: pointer;
            position: relative;
            border: 1px solid #f3f4f6;
            padding: 4px;
            box-sizing: border-box;
        }
        .calendar-day-content {
            width: 100%;
            height: 100%;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            border-radius: 6px;
            padding: 8px 4px;
            box-sizing: border-box;
            transition: all 0.2s;
        }
        .calendar-day-content:hover {
            transform: translateY(-1px);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .calendar-day-empty {
            background: #fafafa;
        }
        .calendar-day-number {
            font-weight: 600;
            color: #1f2937;
            font-size: 1rem;
            line-height: 1;
            margin-bottom: 4px;
        }
        .calendar-day-indicator {
            font-size: 0.75rem;
            margin-top: 2px;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 2px;
        }
        /* 過去の学習データ（緑系） */
        .study-level-0 { background: #f9fafb; }
        .study-level-1 { background: #d1fae5; }
        .study-level-2 { background: #6ee7b7; }
        .study-level-3 { background: #34d399; }
        .study-level-4 { background: #10b981; color: white; }
        
        /* 未来の予定（青系） */
        .future-plan { 
            background: #eff6ff; 
            box-shadow: inset 0 0 0 2px #3b82f6;
        }
        .future-no-plan { background: #f9fafb; }
        
        /* 試験日（赤系） */
        .exam-date { 
            background: linear-gradient(135deg, #fecaca 0%, #ef4444 100%);
            box-shadow: inset 0 0 0 3px #dc2626;
            color: white;
            font-weight: 900;
            position: relative;
            overflow: hidden;
        }
        .exam-badge {
            position: absolute;
            top: 0;
            right: 0;
            background: #dc2626;
            color: white;
            font-size: 0.55rem;
            padding: 1px 4px;
            border-bottom-left-radius: 4px;
            font-weight: 700;
            line-height: 1.2;
        }
        </style>
        """
        
        # HTMLカレンダーを生成
        month_cal = cal.monthcalendar(year, month)
        if st.session_state.language == "English":
             # Use standard English format: "December 2025"
             month_name = datetime(year, month, 1).strftime("%B %Y")
        else:
             month_name = t("month_format").format(year, month)
             
        today = datetime.today().date()
        
        html = f'''
        <div class="calendar-single">
            <div class="calendar-header">
                <div class="calendar-title">{month_name}</div>
                <div class="calendar-nav">
                    <!-- Navigation buttons handled by Streamlit buttons outside HTML -->
                </div>
            </div>
            <table class="calendar-table">
                <tr>
        '''
        
        # 曜日ヘッダー
        weekdays = t("weekdays")
        for wd in weekdays:
            html += f'<th class="calendar-weekday">{wd}</th>'
        html += "</tr>"
        
        # 各週
        for week in month_cal:
            html += "<tr>"
            for day in week:
                if day == 0:
                    # 空白セル
                    html += '<td class="calendar-day"><div class="calendar-day-content calendar-day-empty"></div></td>'
                else:
                    date = datetime(year, month, day).date()
                    
                    # 試験日かチェック
                    is_exam_date = (exam_date and date == exam_date)
                    
                    # 過去 vs 未来
                    is_past = date < today
                    is_today = date == today
                    is_future = date > today
                    
                    tooltip = ""
                    css_class = ""
                    indicator = ""
                    badge = "" # Initialize badge
                    
                    if is_exam_date:
                        # 試験日
                        css_class = "exam-date"
                        tooltip = f"{date.strftime(t('date_format'))}: 🎯{t('exam_date')}"
                        badge = f'<span class="exam-badge">{t("exam_date")}</span>'
                    elif is_past or is_today:
                        # 過去/今日 - 学習データを表示
                        if date in daily_stats_dict:
                            study_time = daily_stats_dict[date]["学習時間"]
                            problems = int(daily_stats_dict[date]["問題数"])
                            accuracy = daily_stats_dict[date]["正答率"] * 100
                            
                            # 色レベルを決定
                            if study_time == 0:
                                level = 0
                            elif study_time <= 30:
                                level = 1
                            elif study_time <= 60:
                                level = 2
                            elif study_time <= 90:
                                level = 3
                            else:
                                level = 4
                            
                            css_class = f"study-level-{level}"
                            tooltip = f"{date.strftime(t('date_format'))}: {problems}{t('questions_unit')}, {t('accuracy_rate')}{accuracy:.0f}%, {int(study_time)}{t('minutes_unit')}"
                            # 絵文字をBootstrap Iconに変更
                            indicator = '<i class="bi bi-check-lg"></i>' if problems > 0 else ""
                        else:
                            css_class = "study-level-0"
                            tooltip = f"{date.strftime(t('date_format'))}: {t('no_data')}"
                        badge = ""
                    else:
                        # 未来 - 週間プランを表示
                        # 日付をキーとして検索
                        plan_count = future_plan_dict.get(date, 0)
                        
                        if plan_count > 0:
                            css_class = "future-plan"
                            tooltip = f"{date.strftime(t('date_format'))}: 📝{t('plan_review')} {plan_count}{t('unit')}"
                            # 絵文字をBootstrap Iconに変更
                            indicator = f'<i class="bi bi-pencil-fill" style="color:#3b82f6; font-size:0.7rem;"></i> <span style="color:#3b82f6;">{plan_count}</span>'
                        else:
                            css_class = "future-no-plan"
                            tooltip = f"{date.strftime(t('date_format'))}: {t('no_change')}"
                        badge = ""
                    
                    html += f'''
                    <td class="calendar-day" title="{tooltip}">
                        <div class="calendar-day-content {css_class}">
                            {badge}
                            <span class="calendar-day-number">{day}</span>
                            <div class="calendar-day-indicator">{indicator}</div>
                        </div>
                    </td>
                    '''
            
            html += "</tr>"
        
        html += '''
            </table>
        </div>
        '''
        
        return css, html
        
    except Exception as e:
        # st is globally imported
        st.error(f"カレンダーヒートマップの生成エラー: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None, None

def generate_detailed_insights(df, current_rate, target_rate, exam_date=None):
    """
    統計分析とパターン認識で、具体的かつ実用的なアドバイスを提供
    """
    if df.empty:
        return []
    
    insights = []
    
    # 1. 学習パターン分析（時間帯・曜日）
    if "日付" in df.columns:
        df["hour"] = pd.to_datetime(df["日付"]).dt.hour
        df["dayofweek"] = pd.to_datetime(df["日付"]).dt.dayofweek
        
        # 時間帯別正答率
        hourly_stats = df.groupby("hour")["ミス"].agg(["sum", "count"])
        hourly_stats["accuracy"] = (hourly_stats["count"] - hourly_stats["sum"]) / hourly_stats["count"]
        
        if len(hourly_stats) >= 2:
            best_hour = hourly_stats["accuracy"].idxmax()
            worst_hour = hourly_stats["accuracy"].idxmin()
            
            if hourly_stats.loc[best_hour, "accuracy"] - hourly_stats.loc[worst_hour, "accuracy"] > 0.15:
                time_label = "朝型" if best_hour < 12 else "午後型" if best_hour < 18 else "夜型"
                insights.append({
                    "category": "学習パターン",
                    "icon": "clock-history",
                    "priority": "high",
                    "message": f"あなたは**{time_label}学習者**です。{best_hour}時台の正答率が最も高いです（{hourly_stats.loc[best_hour, 'accuracy']:.1%}）。重要な学習はこの時間帯に集中させましょう。"
                })
    
    # 2. 弱点の具体的指摘
    unit_stats = df.groupby("単元")["ミス"].agg(["sum", "count"])
    unit_stats["accuracy"] = (unit_stats["count"] - unit_stats["sum"]) / unit_stats["count"]
    unit_stats = unit_stats[unit_stats["count"] >= 3]  # 3問以上のデータがある単元のみ
    
    if not unit_stats.empty:
        weak_units = unit_stats[unit_stats["accuracy"] < 0.5].sort_values("accuracy")
        
        if not weak_units.empty:
            worst_unit = weak_units.index[0]
            worst_accuracy = weak_units.iloc[0]["accuracy"]
            
            # 弱点単元へのアドバイス
            unit_advice = {
                "推論": t("advice_inference"),
                "計算・文章題": t("advice_calculation"),
                "英語": t("advice_english")
            }
            
            advice = unit_advice.get(worst_unit, t("advice_default"))
            
            insights.append({
                "category": t("cat_weakness"),
                "icon": "exclamation-triangle",
                "priority": "high",
                "message": t("insight_weakness_msg").format(dt(worst_unit), worst_accuracy, advice)
            })
    
    # 3. ペース分析
    if exam_date:
        today = datetime.today().date()
        days_left = (exam_date - today).days
        
        if days_left > 0:
            gap = target_rate - current_rate
            required_daily_improvement = gap / days_left if days_left > 0 else 0
            
            if gap > 0.2 and days_left < 30:
                insights.append({
                    "category": t("cat_progress"),
                    "icon": "speedometer",
                    "priority": "urgent",
                    "message": t("insight_urgent_warning").format(days_left=days_left, gap=gap*100, required_daily_improvement=required_daily_improvement*100)
                })
            elif gap > 0 and days_left >= 30:
                insights.append({
                    "category": t("cat_progress"),
                    "icon": "graph-up",
                    "priority": "medium",
                    "message": t("insight_on_track").format(days_left)
                })
            elif gap <= 0:
                insights.append({
                    "category": t("cat_progress"),
                    "icon": "trophy",
                    "priority": "low",
                    "message": t("insight_goal_achieved")
                })
    
    # 4. 比較分析（直近1週間 vs 前週）
    if "日付" in df.columns and len(df) >= 10:
        df["date_obj"] = pd.to_datetime(df["日付"]).dt.date
        today = datetime.today().date()
        week_ago = today - timedelta(days=7)
        two_weeks_ago = today - timedelta(days=14)
        
        this_week = df[df["date_obj"] >= week_ago]
        last_week = df[(df["date_obj"] >= two_weeks_ago) & (df["date_obj"] < week_ago)]
        
        if not this_week.empty and not last_week.empty:
            this_week_rate = (this_week["正誤"] == "〇").sum() / len(this_week)
            last_week_rate = (last_week["正誤"] == "〇").sum() / len(last_week)
            improvement = this_week_rate - last_week_rate
            
            if improvement > 0.05:
                insights.append({
                    "category": t("cat_growth"),
                    "icon": "arrow-up-circle",
                    "priority": "medium",
                    "message": t("insight_growth").format(improvement*100)
                })
            elif improvement < -0.05:
                insights.append({
                    "category": t("cat_growth"),
                    "icon": "arrow-down-circle",
                    "priority": "medium",
                    "message": t("insight_decline").format(abs(improvement)*100)
                })
    
    # 5. 時間管理分析
    if "解答時間(秒)" in df.columns and "目標時間" in df.columns:
        time_excess = (df["解答時間(秒)"] - df["目標時間"]).mean()
        
        if time_excess > 10:
            insights.append({
                "category": t("cat_time"),
                "icon": "hourglass-split",
                "priority": "medium",
                "message": t("insight_time_over").format(time_excess=time_excess)
            })
        elif time_excess < -5:
            insights.append({
                "category": t("cat_time"),
                "icon": "lightning",
                "priority": "low",
                "message": t("insight_time_good")
            })
    
    return insights

def generate_roadmap(exam_date, current_rate, target_rate):
    """
    試験日からの逆算ロードマップ（ガントチャート）生成
    """
    if exam_date is None:
        return None
    
    today = datetime.today().date()
    days_left = (exam_date - today).days
    
    if days_left <= 0:
        return None
        
    # フェーズ計算ロジック
    # 基礎固め: 全体の40% (進捗が遅れていれば50%に拡大)
    # 応用演習: 全体の40%
    # 直前対策: 全体の20%
    
    base_ratio = 0.4
    if current_rate < target_rate - 0.2: # 目標より20%以上低い場合
        base_ratio = 0.5 # 基礎期間を延長
        
    base_days = int(days_left * base_ratio)
    practice_days = int(days_left * (0.8 - base_ratio))
    final_days = days_left - base_days - practice_days
    
    # データフレーム作成
    data = [
        dict(Task=t("timeline_foundation"), Start=today, Finish=today + timedelta(days=base_days), Phase="Foundation"),
        dict(Task=t("timeline_applied"), Start=today + timedelta(days=base_days), Finish=today + timedelta(days=base_days + practice_days), Phase="Practice"),
        dict(Task=t("timeline_final"), Start=today + timedelta(days=base_days + practice_days), Finish=exam_date, Phase="Final")
    ]
    
    df_gantt = pd.DataFrame(data)
    
    # Plotly Expressでガントチャート作成
    fig = px.timeline(df_gantt, x_start="Start", x_end="Finish", y="Task", color="Phase",
                      color_discrete_map={"Foundation": "#60A5FA", "Practice": "#34D399", "Final": "#F87171"},
                      height=150) # 高さを抑える
    
    fig.update_yaxes(autorange="reversed", title=None)
    fig.update_xaxes(title=None, tickformat="%m/%d")
    
    # レイアウト調整
    fig.update_layout(
        margin=dict(l=10, r=10, t=10, b=10),
        showlegend=False,
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(size=12, color="#374151"),
        bargap=0.2
    )
    
    return fig


def generate_study_roadmap_detailed(df, df_master):
    """
    難易度別学習ロードマップの生成
    基礎(低)→標準(中)→応用(高)の順で習熟度を可視化し、次のステップを提案
    """
    if df.empty or df_master.empty:
        return None, None, None
    
    # 難易度列が存在するかチェック（なければデフォルト補完）
    if "難易度" not in df_master.columns:
        # st.warning("マスターデータに「難易度」列がありません。全て「中」として扱います。")
        df_master = df_master.copy()
        df_master["難易度"] = "中"
    
    try:
        # DFとマスタをマージして難易度情報を取得
        if "難易度" in df.columns:
            df_merged = df.copy()
        else:
            df_merged = df.merge(df_master[["問題ID", "難易度", "科目", "単元"]], on="問題ID", how="left")
        
        # 難易度列がNaNの行を除外
        df_merged = df_merged[df_merged["難易度"].notna()]
        
        if df_merged.empty:
            return None, None, None
        
        # 難易度別の統計を計算
        difficulty_stats = {}
        for diff in ["低", "中", "高"]:
            diff_data = df_merged[df_merged["難易度"] == diff]
            if not diff_data.empty:
                total = len(diff_data)
                correct = (diff_data["正誤"] == "〇").sum()
                accuracy = correct / total if total > 0 else 0
                
                # マスタデータから該当難易度の総問題数を取得
                master_diff = df_master[df_master["難易度"] == diff]
                total_problems_in_master = len(master_diff)
                coverage = (len(diff_data["問題ID"].unique()) / total_problems_in_master * 100) if total_problems_in_master > 0 else 0
                
                # その難易度の主な単元（問題数が多い順トップ5）
                top_units = master_diff["単元"].value_counts().head(5).index.tolist()
                
                difficulty_stats[diff] = {
                    "solved": len(diff_data["問題ID"].unique()),
                    "total": total_problems_in_master,
                    "accuracy": accuracy,
                    "coverage": coverage,
                    "attempts": total,
                    "units": top_units
                }
            else:
                # データがない場合
                master_diff = df_master[df_master["難易度"] == diff]
                total_problems_in_master = len(master_diff)
                top_units = master_diff["単元"].value_counts().head(5).index.tolist()
                
                difficulty_stats[diff] = {
                    "solved": 0,
                    "total": total_problems_in_master,
                    "accuracy": 0,
                    "coverage": 0,
                    "attempts": 0,
                    "units": top_units
                }
        
        # 現在のフェーズを判定
        current_phase = "基礎固め"
        next_recommendations = []
        
        # 基礎(低)が80%以上の正答率かつ70%以上のカバレッジなら標準へ
        if difficulty_stats["低"]["accuracy"] >= 0.8 and difficulty_stats["低"]["coverage"] >= 70:
            # 標準(中)が80%以上の正答率かつ70%以上のカバレッジなら応用へ
            if difficulty_stats["中"]["accuracy"] >= 0.8 and difficulty_stats["中"]["coverage"] >= 70:
                current_phase = "応用演習"
                next_recommendations = [
                    t("rec_continue_advanced"),
                    t("rec_aim_high_accuracy"),
                    t("rec_reduce_time")
                ]
            else:
                current_phase = "標準演習"
                # 未着手の標準問題を推薦
                unsolved_medium = df_master[
                    (df_master["難易度"] == "中") & 
                    (~df_master["問題ID"].isin(df["問題ID"].unique()))
                ]
                if not unsolved_medium.empty:
                    top_units = unsolved_medium["単元"].value_counts().head(3).index.tolist()
                    next_recommendations = [
                        t("rec_next_challenge").format(dt(top_units[0])),
                        t("rec_aim_standard_80"),
                        t("rec_current_coverage").format(difficulty_stats['中']['coverage'])
                    ]
                else:
                    next_recommendations = [
                        t("rec_review_standard"),
                        t("rec_aim_stable_80")
                    ]
        else:
            current_phase = "基礎固め"
            # 未着手の基礎問題を推薦
            unsolved_low = df_master[
                (df_master["難易度"] == "低") & 
                (~df_master["問題ID"].isin(df["問題ID"].unique()))
            ]
            if not unsolved_low.empty:
                top_units = unsolved_low["単元"].value_counts().head(3).index.tolist()
                next_recommendations = [
                    t("rec_start_basic").format(dt(top_units[0])),
                    t("rec_aim_basic_80"),
                    t("rec_current_coverage").format(difficulty_stats['低']['coverage'])
                ]
            else:
                next_recommendations = [
                    t("rec_review_basic"),
                    t("rec_aim_stable_80")
                ]
        
        # ビジュアライゼーション用データ作成
        roadmap_data = {
            "phase": [t("phase_foundation"), t("phase_standard"), t("phase_advanced")],
            "progress": [
                difficulty_stats["低"]["coverage"],
                difficulty_stats["中"]["coverage"],
                difficulty_stats["高"]["coverage"]
            ],
            "units": [
                [dt(u) for u in difficulty_stats["低"]["units"]],
                [dt(u) for u in difficulty_stats["中"]["units"]],
                [dt(u) for u in difficulty_stats["高"]["units"]]
            ],
            "accuracy": [
                difficulty_stats["低"]["accuracy"] * 100,
                difficulty_stats["中"]["accuracy"] * 100,
                difficulty_stats["高"]["accuracy"] * 100
            ],
            "status": [
                t("status_completed") if difficulty_stats["低"]["accuracy"] >= 0.8 and difficulty_stats["低"]["coverage"] >= 70 else t("status_in_progress") if difficulty_stats["低"]["attempts"] > 0 else t("status_not_started"),
                t("status_completed") if difficulty_stats["中"]["accuracy"] >= 0.8 and difficulty_stats["中"]["coverage"] >= 70 else t("status_in_progress") if difficulty_stats["中"]["attempts"] > 0 else t("status_not_started"),
                t("status_completed") if difficulty_stats["高"]["accuracy"] >= 0.8 and difficulty_stats["高"]["coverage"] >= 70 else t("status_in_progress") if difficulty_stats["高"]["attempts"] > 0 else t("status_not_started")
            ]
        }
        
        return roadmap_data, current_phase, next_recommendations
        
    except Exception as e:
        st.error(t("roadmap_error").format(e))
        return None, None, None

def generate_stacked_bar_chart(df):
    """
    学習フローの積み上げ棒グラフ生成
    単元ごとの正解・不正解数を積み上げ棒グラフで表示
    """
    if df.empty or len(df) < 5:
        return None
    
    # データ準備
    df_bar = df.copy()
    df_bar["正誤ラベル"] = df_bar["正誤"].apply(lambda x: t("correct") if x == "〇" else t("incorrect"))
    df_bar["単元ラベル"] = df_bar["単元"].apply(dt)
    
    # 集計: 単元・正誤ごとの件数
    bar_data = df_bar.groupby(["単元ラベル", "正誤ラベル"]).size().reset_index(name="count")
    
    # 合計件数でソート（多い順）
    total_counts = bar_data.groupby("単元ラベル")["count"].sum().sort_values(ascending=True)
    bar_data["単元ラベル"] = pd.Categorical(bar_data["単元ラベル"], categories=total_counts.index, ordered=True)
    bar_data = bar_data.sort_values("単元ラベル")
    
    # 積み上げ棒グラフ作成
    fig = px.bar(
        bar_data,
        y="単元ラベル",
        x="count",
        color="正誤ラベル",
        orientation='h',
        color_discrete_map={
            t("correct"): "rgba(16, 185, 129, 0.8)",   # Green with opacity
            t("incorrect"): "rgba(239, 68, 68, 0.8)"   # Red with opacity
        },
        text="count"
    )
    
    fig.update_traces(
        textposition='inside', 
        textfont_color='white',
        hovertemplate='%{y}<br>%{data.name}: %{x}問<extra></extra>'
    )
    
    fig.update_layout(
        title=dict(
            text=t("learning_flow_visualization"),
            font=dict(size=18, color="#111827", weight="bold"),
            x=0.5,
            xanchor="center"
        ),
        xaxis_title=None, # Remove redundant title
        yaxis_title=None,
        barmode='stack',
        height=max(400, len(total_counts) * 30), # Increase height per bar
        margin=dict(l=10, r=10, t=50, b=10),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=None),
        xaxis=dict(
            showgrid=True, 
            gridcolor='rgba(0,0,0,0.05)',
            zeroline=False,
            tickfont=dict(color="#4b5563")
        ),
        yaxis=dict(
            tickfont=dict(color="#1f2937", size=13)
        )
    )
    
    return fig

def generate_weekly_report(df):
    """
    週報レポート生成（過去7日間の学習サマリー）
    """
    if df.empty:
        return t("report_no_data")
    
    today = datetime.today().date()
    week_ago = today - timedelta(days=7)
    
    # 過去7日間のデータ
    df["date_obj"] = pd.to_datetime(df["日付"]).dt.date
    df_week = df[df["date_obj"] >= week_ago].copy()
    
    if df_week.empty:
        return t("report_no_week_data")
    
    # 集計
    total_problems = len(df_week)
    total_time = df_week["学習投入時間(分)"].sum() if "学習投入時間(分)" in df_week.columns else 0
    accuracy = (1 - df_week["ミス"].mean()) * 100
    
    # 最も頑張った単元
    top_unit = df_week.groupby("単元").size().idxmax() if not df_week.empty else "N/A"
    top_count = df_week.groupby("単元").size().max() if not df_week.empty else 0
    
    # 継続日数
    study_days = df_week["date_obj"].nunique()
    
    report = f"""
### <i class="bi bi-bar-chart-fill"></i> **{t("report_title").format(st.session_state.current_user)}**
{t("report_period").format(week_ago.strftime('%Y/%m/%d'), today.strftime('%Y/%m/%d'))}

---

### <i class="bi bi-graph-up"></i> {t("report_achievements")}
{t("report_study_days").format(study_days)}
{t("report_total_problems").format(total_problems)}
{t("report_total_time").format(total_time, total_time/60)}
{t("report_avg_accuracy").format(accuracy)}

### <i class="bi bi-trophy-fill"></i> {t("report_top_unit_title")}
{t("report_top_unit_desc").format(dt(top_unit), top_count)}

### <i class="bi bi-chat-quote-fill"></i> {t("report_ai_comment_title")}
"""
    
    # 簡易的な総評ロジック
    if accuracy >= 80:
        report += t("report_comment_excellent")
    elif accuracy >= 60:
        report += t("report_comment_good")
    else:
        report += t("report_comment_basic")
    
    report += f"\n\n### <i class='bi bi-bullseye'></i> {t('report_next_goal_title')}\n{t('report_next_goal_desc').format(min(100, accuracy + 5))}\n"
    
    return report

def predict_with_prophet(df, target_rate, exam_date):
    """
    Prophet時系列予測 - より精密な正答率予測
    トレンド + 季節性を考慮した予測を提供
    """
    try:
        from prophet import Prophet
    except ImportError:
        return None, t("prophet_not_installed")
    
    if df.empty or len(df) < 10:
        return None, t("prophet_min_data")
    
    if exam_date is None:
        return None, t("prophet_no_exam_date")
    
    # データ準備
    df_prophet = df.copy()
    df_prophet["ds"] = pd.to_datetime(df_prophet["日付"])
    
    # 日別正答率を計算
    daily_accuracy = df_prophet.groupby("ds").apply(
        lambda x: (x["正誤"] == "〇").sum() / len(x)
    ).reset_index()
    daily_accuracy.columns = ["ds", "y"]
    
    if len(daily_accuracy) < 2:
        return None, "予測には最低2日分のデータが必要です"
    
    # Prophetモデル構築
    model = Prophet(
        daily_seasonality=False,
        weekly_seasonality=True if len(daily_accuracy) >= 7 else False,
        yearly_seasonality=False,
        interval_width=0.8  # 80%信頼区間
    )
    
    model.fit(daily_accuracy)
    
    # 未来予測（試験日まで）
    future_dates = model.make_future_dataframe(periods=(exam_date - datetime.today().date()).days)
    forecast = model.predict(future_dates)
    
    # 試験日の予測値
    exam_datetime = pd.Timestamp(exam_date)
    exam_prediction = forecast[forecast["ds"] == exam_datetime]
    
    if exam_prediction.empty:
        # 試験日がデータ範囲外の場合、最も近い日付を使用
        exam_prediction = forecast.iloc[-1]
        predicted_rate = exam_prediction["yhat"]
    else:
        predicted_rate = exam_prediction["yhat"].values[0]
    
    # 予測値をグラフ用に整形
    forecast_display = forecast[["ds", "yhat", "yhat_lower", "yhat_upper"]].copy()
    forecast_display.columns = ["日付", "予測正答率", "下限", "上限"]
    
    return {
        "forecast": forecast_display,
        "predicted_rate": predicted_rate,
        "model": model,
        "actual_data": daily_accuracy
    }, None

def generate_pdf_report(report_text, user_name, df=None):
    """
    週報レポートをPDF化（日本語対応・グラフ付き）
    """
    try:
        from fpdf import FPDF
        import io
        import matplotlib.pyplot as plt
        import tempfile
        
        class PDF(FPDF):
            def header(self):
                # ヘッダー
                # 日本語フォントが読み込まれていればそれを使う、なければArial
                if 'jp' in self.font_files:
                    self.set_font('jp', 'B', 16)
                else:
                    self.set_font('Arial', 'B', 16)
                self.cell(0, 10, 'SPI Learning Report', 0, 1, 'C')
                self.ln(5)
            
            def footer(self):
                # フッター
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
        
        pdf = PDF()
        
        # フォント読み込み（NotoSansJP-Regular.ttf）
        font_path = "fonts/NotoSansJP-Regular.ttf"
        if os.path.exists(font_path):
            pdf.add_font('jp', '', font_path, uni=True)
            pdf.add_font('jp', 'B', font_path, uni=True) # Boldも同じフォントで代用
            font_family = 'jp'
        else:
            font_family = 'Arial' # フォールバック
            
        pdf.add_page()
        pdf.set_font(font_family, size=10)
        
        # --- グラフ生成と埋め込み ---
        if df is not None and not df.empty:
            try:
                # 科目別正答率グラフ
                plt.figure(figsize=(6, 4))
                subject_acc = df.groupby("科目")["ミス"].agg(["sum", "count"]).reset_index()
                subject_acc["accuracy"] = (subject_acc["count"] - subject_acc["sum"]) / subject_acc["count"]
                
                # 日本語フォント設定（matplotlib用）
                # 環境によっては豆腐になるため、英語ラベルにするか、フォントパスを指定する
                # ここでは簡易的に英語ラベルを使用
                plt.bar(subject_acc["科目"], subject_acc["accuracy"], color="#3B82F6")
                plt.title("Subject Accuracy")
                plt.ylim(0, 1)
                plt.ylabel("Accuracy")
                
                # 一時ファイルに保存
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
                    plt.savefig(tmp_file.name, format="png", dpi=100)
                    tmp_path = tmp_file.name
                
                # PDFに追加
                pdf.image(tmp_path, x=10, y=30, w=100)
                pdf.ln(80) # 画像分スペースを空ける
                
                # 後始末
                os.remove(tmp_path)
            except Exception as e:
                pdf.multi_cell(0, 5, f"[Graph Error: {e}]")
                pdf.ln(5)

        # レポート本文
        clean_text = report_text.replace("**", "").replace("###", "").replace("##", "").replace("*", "")
        
        for line in clean_text.split("\n"):
            if line.strip():
                try:
                    pdf.multi_cell(0, 6, line)
                except:
                    pdf.multi_cell(0, 6, "[Text Error]")
            else:
                pdf.ln(3)
        
        # バイナリデータとして返す
        pdf_output = io.BytesIO()
        pdf_data = pdf.output(dest='S').encode('latin-1')
        pdf_output.write(pdf_data)
        pdf_output.seek(0)
        
        return pdf_output
        
    except ImportError:
        return None
    except Exception as e:
        st.error(f"PDF Generation Error: {e}")
        return None

def generate_excel_report(df, user_name):
    """
    学習ログと統計をExcel形式で出力
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # シート1: 生データ
        ws1 = wb.active
        ws1.title = "学習ログ"
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # データフレームをExcelに書き込み
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # ヘッダー行
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
        
        # 列幅自動調整
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # シート2: 統計サマリー
        ws2 = wb.create_sheet("統計サマリー")
        
        if not df.empty:
            # 基本統計
            total_problems = len(df)
            correct_count = (df["正誤"] == "〇").sum()
            accuracy = correct_count / total_problems
            
            stats_data = [
                ["指標", "値"],
                ["総問題数", total_problems],
                ["正解数", correct_count],
                ["正答率", f"{accuracy:.1%}"],
                ["平均解答時間", f"{df['解答時間(秒)'].mean():.1f}秒"],
                ["総学習時間", f"{df['学習投入時間(分)'].sum():.0f}分"]
            ]
            
            for r_idx, row in enumerate(stats_data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
            
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 15
        
        # バイナリデータとして返す
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        return excel_output
        
    except ImportError:
        return None

# ページ設定
st.set_page_config(
    page_title="SPI対策 Dashboard",
    page_icon="�",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Google Login Logic ---
if "current_user" not in st.session_state:
    st.session_state.current_user = None

if not st.session_state.current_user:
    # Try to login automatically
    creds, error = get_credentials()
    if creds:
        user_info, error = get_user_info(creds)
        if user_info:
            st.session_state.current_user = user_info.get('email')
            st.session_state.user_name = user_info.get('name')
            
            # Load user settings
            if "sheets_manager" not in st.session_state:
                st.session_state.sheets_manager = GoogleSheetsManager()
            
            settings, err = st.session_state.sheets_manager.load_user_settings(st.session_state.current_user)
            if settings:
                st.session_state.company_name = settings.get("company_name", "")
                st.session_state.target_rate_user = settings.get("target_rate_user", 80)
                st.session_state.daily_study_time = settings.get("daily_study_time", 60)
                st.session_state.time_policy = settings.get("time_policy", "標準")
                st.session_state.exam_date = settings.get("exam_date")
            
            st.rerun()
        else:
            if "invalid_scope" in str(error):
                st.error("認証スコープエラー: 保存されているトークンの権限が不足しています。")
                st.warning("""
                **【重要】Streamlit CloudのSecretsを更新してください**
                
                新しい機能（Googleログイン）には、追加の権限（メールアドレスの取得）が必要です。
                以下の手順でトークンを再生成し、Secretsを更新してください：
                
                1. ローカル環境でアプリを実行し、ログインし直す。
                2. 生成された `token.json` の中身をコピーする。
                3. Streamlit Cloudのアプリ設定画面 > Secrets に移動する。
                4. `[token]` セクションの中身を、新しい `token.json` の内容で上書きする。
                """)
                st.stop()
            else:
                st.error(f"Login Failed: {error}")
                if st.button("Retry Login"):
                    if os.path.exists('token.json'):
                        os.remove('token.json')
                    st.rerun()
                st.stop()
    else:
        st.info("Logging in...")
        # get_credentials should have triggered the flow. If it returned None/Error without flow, show error.
        st.error(f"Authentication Error: {error}")
        if st.button("Retry"):
             st.rerun()
        st.stop()

# --- Bootstrap Icons & Custom CSS ---
st.markdown("""
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
<style>
/* アイコンバッジ */
.icon-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 32px;
    height: 32px;
    border-radius: 8px;
    background-color: #eff6ff; /* 薄い青 */
    color: #3b82f6; /* 青 */
    margin-right: 10px;
    font-size: 1.1rem;
    flex-shrink: 0;
}
/* チャートタイトル用ラッパー */
.chart-header {
    display: flex;
    align-items: center;
    margin-bottom: 12px;
    font-weight: 700;
    font-size: 1.1rem;
    color: #1f2937;
}
</style>
""", unsafe_allow_html=True)

# ===== テーマ定義 & カラー設定 =====
if "theme" not in st.session_state:
    st.session_state.theme = "Blue"

THEMES = {
    "Blue": {
        "PRIMARY": "#3B82F6", "ACCENT": "#F97316", "SUCCESS": "#10B981", 
        "WARNING": "#F59E0B", "DANGER": "#EF4444", "NEUTRAL": "#6B7280", "BACKGROUND": "#F8FAFC"
    },
    "Green": {
        "PRIMARY": "#059669", "ACCENT": "#D97706", "SUCCESS": "#3B82F6", 
        "WARNING": "#F59E0B", "DANGER": "#EF4444", "NEUTRAL": "#6B7280", "BACKGROUND": "#F0FDF4"
    },
    "Orange": {
        "PRIMARY": "#EA580C", "ACCENT": "#0284C7", "SUCCESS": "#10B981", 
        "WARNING": "#F59E0B", "DANGER": "#EF4444", "NEUTRAL": "#6B7280", "BACKGROUND": "#FFF7ED"
    },
    "Dark": {
        "PRIMARY": "#60A5FA", "ACCENT": "#FB923C", "SUCCESS": "#34D399", 
        "WARNING": "#FBBF24", "DANGER": "#F87171", "NEUTRAL": "#9CA3AF", "BACKGROUND": "#0F172A"
    },
}

# テーマ取得（キーエラー対策）
current_theme = st.session_state.theme
if current_theme not in THEMES:
    current_theme = "Blue"
    st.session_state.theme = "Blue"

tm = THEMES[current_theme]
PRIMARY = tm["PRIMARY"]
ACCENT = tm["ACCENT"]
SUCCESS = tm["SUCCESS"]
WARNING = tm["WARNING"]
DANGER = tm["DANGER"]
BACKGROUND = tm["BACKGROUND"]
DANGER = tm["DANGER"]
NEUTRAL = tm["NEUTRAL"]
BACKGROUND = tm["BACKGROUND"]

# ===== 問題マスタ（全30問） =====
MASTER_COLUMNS = ["問題ID", "科目", "ジャンル", "単元", "目標解答時間(秒)", "目標正答率(%)", "難易度", "出題頻度(重み)"]
DEFAULT_MASTER_ROWS = [
    ["N-A01", "非言語", "推論", "集合の推論 (ベン図)", 120, 85, "高", 4],
    ["N-A02", "非言語", "推論", "論理的な推論 (真偽・順序)", 100, 80, "高", 5],
    ["N-A03", "非言語", "推論", "対戦・リーグ戦の推論", 150, 75, "高", 5],
    ["N-A04", "非言語", "推論", "命題・三段論法", 90, 90, "低", 3],
    ["N-A05", "非言語", "推論", "領域 (座標平面)", 180, 65, "高", 3],
    ["N-A06", "非言語", "推論", "物の流れ", 160, 70, "高", 4],
    ["N-B01", "非言語", "計算・文章題", "比と割合の計算", 60, 95, "低", 5],
    ["N-B02", "非言語", "計算・文章題", "濃度算", 100, 80, "中", 5],
    ["N-B03", "非言語", "計算・文章題", "割引・割増計算", 90, 85, "中", 3],
    ["N-B04", "非言語", "計算・文章題", "損益算", 110, 90, "中", 5],
    ["N-B05", "非言語", "計算・文章題", "仕事算", 90, 80, "中", 4],
    ["N-B06", "非言語", "計算・文章題", "速度算", 130, 75, "高", 5],
    ["N-B07", "非言語", "計算・文章題", "料金の割引", 100, 70, "中", 5],
    ["N-C01", "非言語", "確率・場合", "確率 (基礎)", 70, 70, "低", 5],
    ["N-C02", "非言語", "確率・場合", "場合の数", 120, 65, "高", 5],
    ["N-D01", "非言語", "図表の読み取り", "グラフ・表の計算", 150, 75, "高", 3],
    ["N-D02", "非言語", "図表の読み取り", "増加率の把握", 140, 80, "中", 4],
    ["N-D03", "非言語", "図表の読み取り", "複数情報の読み取り", 180, 70, "高", 3],
    ["N-E01", "非言語", "特殊算・その他", "植木算・年齢算", 90, 85, "低", 3],
    ["N-E02", "非言語", "特殊算・その他", "集合の計算", 100, 80, "中", 3],
    ["N-E03", "非言語", "特殊算・その他", "分割払い", 110, 75, "中", 3],
    ["N-E04", "非言語", "特殊算・その他", "不定方程式", 120, 60, "低", 1],
    ["L-A01", "言語", "語彙知識", "二語の関係", 15, 95, "低", 4],
    ["L-A02", "言語", "語彙知識", "熟語の成り立ち", 20, 90, "中", 4],
    ["L-A03", "言語", "語彙知識", "語句の定義", 25, 85, "中", 5],
    ["L-B01", "言語", "文法・表現", "語句の用法", 30, 80, "中", 4],
    ["L-B02", "言語", "文法・表現", "空欄補充", 40, 75, "中", 4],
    ["L-B03", "言語", "文法・表現", "文の並べ替え", 100, 70, "高", 5],
    ["L-C01", "言語", "文章読解", "長文読解", 480, 70, "高", 5],
    ["L-C02", "言語", "文章読解", "論理的読解", 180, 65, "高", 4],
]
df_master_default = pd.DataFrame(DEFAULT_MASTER_ROWS, columns=MASTER_COLUMNS)

# ===== セッション初期化 =====
if "df_log_manual" not in st.session_state:
    st.session_state.df_log_manual = pd.DataFrame(columns=["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"])
if "target_rate_user" not in st.session_state:
    st.session_state.target_rate_user = 80
if "company_name" not in st.session_state:
    st.session_state.company_name = ""
if "time_policy" not in st.session_state:
    st.session_state.time_policy = "標準"
if "subj" not in st.session_state:
    st.session_state.subj = None
if "gen" not in st.session_state:
    st.session_state.gen = None
if "uni" not in st.session_state:
    st.session_state.uni = None
if "keep_input_open" not in st.session_state:
    st.session_state.keep_input_open = True
if "expander_open" not in st.session_state:
    st.session_state.expander_open = st.session_state.keep_input_open
if "exam_date" not in st.session_state:
    st.session_state.exam_date = None
if "language" not in st.session_state:
    st.session_state.language = "日本語"
# if "current_user" not in st.session_state:
#     st.session_state.current_user = "デフォルトユーザー"
if "user_data_dir" not in st.session_state:
    st.session_state.user_data_dir = "user_data"
if "daily_study_time" not in st.session_state:
    st.session_state.daily_study_time = 60
if "plan_completion" not in st.session_state:
    st.session_state.plan_completion = {}
if "df_notes" not in st.session_state:
    st.session_state.df_notes = pd.DataFrame(columns=["問題ID", "メモ", "登録日時"])
if "display_mode" not in st.session_state:
    st.session_state.display_mode = "システム設定"

# ===== 高品質CSS (Glassmorphism & Modern UI) =====

# ダークモード用スタイル定義
dark_css = """
    /* ルート変数定義 */
    /* ルート変数定義 (Streamlit変数の強制オーバーライド) */
    :root {
        /* カスタム変数 */
        --primary: #60a5fa;
        --accent: #fb923c;
        --success: #34d399;
        --warning: #fbbf24;
        --danger: #f87171;
        --neutral: #9ca3af;
        --background: #0f172a;
        --surface: #1e293b;
        --text-primary: #f1f5f9;
        --text-secondary: #cbd5e1;
        --border-color: #334155;

        /* Streamlit標準変数のオーバーライド */
        --primary-color: #60a5fa !important;
        --background-color: #0f172a !important;
        --secondary-background-color: #1e293b !important;
        --text-color: #f1f5f9 !important;
        --font: "sans serif" !important;
    }

    /* アプリ全体の背景とテキスト */
    html, body, .stApp {
        background-color: #0f172a !important;
        color: #f1f5f9 !important;
    }

    /* サイドバーの背景 */
    [data-testid="stSidebar"], [data-testid="stSidebar"] > div {
        background-color: #1e293b !important;
    }

    /* ============================================
       ダークモード修正 (完結編・最終調整 V4)
       ============================================ */
    
    /* 0. ブラウザネイティブのダークモード有効化 */
    :root {
        color-scheme: dark;
    }
    
    /* 1. アプリ全体とサイドバー */
    .stApp {
        background-color: #0f172a !important;
        color: #f1f5f9 !important;
    }
    
    [data-testid="stSidebar"] {
        background-color: #1e293b !important;
        border-right: 1px solid #334155 !important;
    }
    
    /* 2. 入力フォームの徹底修正（属性セレクタ使用） */
    .stApp input[type="text"],
    .stApp input[type="number"],
    .stApp input[type="date"],
    .stApp input[type="password"],
    .stApp input[type="email"],
    .stApp textarea,
    .stApp select {
        background-color: #334155 !important;
        color: #f1f5f9 !important;
        border-color: #475569 !important;
        caret-color: #f1f5f9 !important;
    }
    
    /* BaseWebコンテナの修正 */
    .stApp div[data-baseweb="input"],
    .stApp div[data-baseweb="base-input"],
    .stApp div[data-baseweb="textarea"],
    .stApp div[data-baseweb="select"] > div {
        background-color: #334155 !important;
        border-color: #475569 !important;
        color: #f1f5f9 !important;
    }

    /* サイドバー入力の強制オーバーライド (base_cssの特異性に対抗) */
    [data-testid="stSidebar"] .stSelectbox > div > div,
    [data-testid="stSidebar"] .stTextInput > div > div > input,
    [data-testid="stSidebar"] .stNumberInput > div > div > input,
    [data-testid="stSidebar"] .stDateInput > div > div > input,
    [data-testid="stSidebar"] .stTextArea > div > div > textarea,
    [data-testid="stSidebar"] [data-baseweb="select"] > div,
    [data-testid="stSidebar"] [data-baseweb="popover"] {
        background-color: #334155 !important;
        border-color: #475569 !important;
        color: #f1f5f9 !important;
    }
    
    [data-testid="stSidebar"] [role="listbox"],
    [data-testid="stSidebar"] [role="option"] {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
    }
    
    /* 3. ボタン（計測開始・停止など）の修正 */
    .stApp button {
        background-color: #334155 !important;
        color: #f1f5f9 !important;
        border-color: #475569 !important;
    }
    
    .stApp button:hover {
        border-color: #60a5fa !important;
        color: #60a5fa !important;
    }
    
    /* Primaryボタン（データ追加など）は赤色を維持 */
    .stApp button[kind="primary"] {
        background-color: #ef4444 !important;
        border-color: #ef4444 !important;
        color: white !important;
    }
    .stApp button[kind="primary"]:hover {
        background-color: #dc2626 !important;
    }
    
    /* 4. 数値入力のステップボタン（+/-）の修正 */
    .stApp [data-baseweb="spin-button-group"] {
        background-color: #334155 !important;
        color: #f1f5f9 !important;
    }
    
    .stApp [data-baseweb="spin-button-group"] > div {
        background-color: #334155 !important;
        color: #f1f5f9 !important;
        border-color: #475569 !important;
    }
    
    /* 5. ファイルアップローダーの修正 */
    .stApp [data-testid="stFileUploaderDropzone"] {
        background-color: #334155 !important;
        border-color: #475569 !important;
        color: #f1f5f9 !important;
    }
    
    .stApp [data-testid="stFileUploaderDropzone"] div,
    .stApp [data-testid="stFileUploaderDropzone"] span,
    .stApp [data-testid="stFileUploaderDropzone"] small {
        color: #f1f5f9 !important;
    }
    
    /* 6. Expander (details/summaryタグ) */
    .stApp details {
        background-color: #1e293b !important;
        border-color: #334155 !important;
        color: #f1f5f9 !important;
        border-radius: 8px !important;
    }
    
    .stApp summary {
        background-color: transparent !important;
        color: #f1f5f9 !important;
    }
    
    .stApp summary:hover {
        color: #60a5fa !important;
    }
    
    .stApp [data-testid="stExpanderDetails"] {
        background-color: transparent !important;
        color: #f1f5f9 !important;
    }
    
    /* 7. プレースホルダー */
    .stApp ::placeholder {
        color: #94a3b8 !important;
        opacity: 0.7 !important;
    }
    
    /* 8. タブバー (SAC/Ant Design 強制オーバーライド - ダークモード) */
    /* 重要: base_cssのライトモードスタイルを確実に上書きするため、transparentではなく実際の暗い色を指定 */
    .stApp .stTabs,
    .stApp .ant-tabs,
    .stApp .ant-tabs-top {
        background-color: transparent !important;
    }
    
    /* これらの要素はbase_cssで白(#ffffff)に設定されているため、暗い色で上書き */
    .stApp .ant-tabs-nav,
    .stApp .ant-tabs-nav-wrap,
    .stApp .ant-tabs-nav-list {
        background-color: #1e293b !important;
        background: #1e293b !important;
        border: 1px solid #334155 !important;
        border-radius: 8px !important;
        padding: 4px !important;
    }

    /* タブのコンテナ */
    .stApp div[data-baseweb="tab-list"],
    .stApp .ant-tabs-nav-operations {
        background-color: #1e293b !important;
        border: 1px solid #334155 !important;
        padding: 4px !important;
        border-radius: 8px !important;
        gap: 4px !important;
    }
    
    /* 個別のタブボタン */
    .stApp button[data-baseweb="tab"],
    .stApp .ant-tabs-tab {
        background-color: transparent !important;
        color: #94a3b8 !important;
        border-radius: 6px !important;
        border: none !important;
        margin: 0 !important;
    }
    
    /* アクティブなタブ */
    .stApp button[data-baseweb="tab"][aria-selected="true"],
    .stApp .ant-tabs-tab-active {
        background-color: #3b82f6 !important;
        color: white !important;
    }
    
    /* タブボタン内のテキスト */
    .stApp .ant-tabs-tab-btn {
        color: inherit !important;
    }


    /* SAC Divider Fix (Ant Design Divider) - Nuclear (No .stApp) */
    html body .ant-divider,
    html body .ant-divider *,
    html body div[class*="ant-divider"],
    html body div[class*="ant-divider"] * {
        border-top-color: #334155 !important;
        color: #f1f5f9 !important;
    }
    html body .ant-divider-inner-text,
    html body .ant-divider-inner-text *,
    html body div[class*="ant-divider-inner-text"] {
        background-color: #0f172a !important;
        color: #f1f5f9 !important;
    }

    /* タブバーの背景色を強力に上書き - Nuclear (No .stApp) */
    html body .ant-tabs-nav,
    html body .ant-tabs-nav *,
    html body div[class*="ant-tabs-nav"],
    html body div[class*="ant-tabs-nav"] * {
        background-color: #1e293b !important;
        background: #1e293b !important;
        border-color: #334155 !important;
    }
    
    html body .ant-tabs-tab,
    html body div[class*="ant-tabs-tab"] {
        background-color: transparent !important;
    }
    
    html body .ant-tabs-tab-active,
    html body div[class*="ant-tabs-tab-active"] {
        background-color: #3b82f6 !important;
        color: white !important;
    }

    /* 水平線 (hr) */
    .stApp hr {
        border-color: #334155 !important;
        opacity: 1 !important;
    }
    
    /* 9. アラート・カード類 */
    .stApp [data-testid="stAlert"] {
        background-color: rgba(30, 41, 59, 0.95) !important;
        border: 1px solid #3b82f6 !important;
        color: #f1f5f9 !important;
    }
    
    .metric-card, .action-card {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%) !important;
        border: 1px solid #334155 !important;
    }
    
    .metric-value, .action-unit, .action-title, .metric-label {
        color: #f1f5f9 !important;
    }
    
    /* 10. ドロップダウンメニュー（ポータル） */
    div[data-baseweb="popover"], div[data-baseweb="menu"], ul[role="listbox"] {
        background-color: #1e293b !important;
        border: 1px solid #475569 !important;
    }
    
    li[role="option"] {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
    }
    
    li[role="option"]:hover, li[role="option"][aria-selected="true"] {
        background-color: #334155 !important;
    }
    
    /* 11. テキストカラー強制 */
    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6, 
    .stApp p, .stApp label, .stApp span, .stApp div, .stApp li {
        color: #f1f5f9 !important;
    }
    
    /* 例外: Primaryボタンのテキスト */
    .stApp button[kind="primary"] span {
        color: white !important;
    }
    
    /* 例外: カレンダーの日付 */
    .stApp div[data-baseweb="calendar"] button {
        color: #f1f5f9 !important;
    }
    .stApp div[data-baseweb="calendar"] button:hover {
        background-color: #3b82f6 !important;
    }
"""

base_css = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Noto+Sans+JP:wght@400;500;700;900&display=swap');

:root {
  --primary: {PRIMARY};
  --accent: {ACCENT};
  --success: {SUCCESS};
  --warning: {WARNING};
  --danger: {DANGER};
  --neutral: {NEUTRAL};
  --background: {BACKGROUND};
}

* {
    font-family: 'Inter', 'Noto Sans JP', sans-serif;
    box-sizing: border-box;
}

.stApp {
    background-color: var(--background);
    background-image: 
        radial-gradient(at 0% 0%, rgba(59, 130, 246, 0.05) 0px, transparent 50%),
        radial-gradient(at 100% 0%, rgba(249, 115, 22, 0.05) 0px, transparent 50%);
}

/* ヘッダー */
.header {
    position: sticky; top: 0; z-index: 20;
    padding: 16px 0;
    margin-bottom: 24px;
}
.title-wrap { display:flex; align-items:center; gap:16px; }
.logo { 
    width:48px; height:48px; border-radius:12px; 
    display:flex; align-items:center; justify-content:center; 
    background: linear-gradient(135deg, var(--primary), #1e40af);
    color:#fff; font-weight:800; font-size: 24px;
    box-shadow: 0 4px 6px -1px rgba(59, 130, 246, 0.3);
}

/* アクションカード */
.action-card {
    background: linear-gradient(135deg, #fff7ed 0%, #ffffff 100%);
    border: 2px solid var(--accent);
    border-radius: 16px;
    padding: 24px;
    box-shadow: 0 10px 15px -3px rgba(249, 115, 22, 0.1);
    display:flex; gap:20px; align-items:flex-start;
    position: relative; overflow: hidden;
}
.action-card::before {
    content: ''; position: absolute; top: 0; right: 0; width: 100px; height: 100px;
    background: var(--accent); opacity: 0.05; border-radius: 0 0 0 100%;
}
.action-icon {
    width:64px; height:64px; border-radius:16px;
    background: var(--accent); color:white;
    display:flex; align-items:center; justify-content:center;
    font-size:28px; flex-shrink:0;
    box-shadow: 0 4px 6px -1px rgba(249, 115, 22, 0.3);
}
.action-content { flex:1; z-index: 1; }
.action-title { color: #1f2937; font-weight:800; font-size:1.1rem; margin:0; }
.priority-badge {
    background: var(--danger); color: white; padding: 4px 12px;
    border-radius: 999px; font-size: 0.75rem; font-weight: 700;
    margin-left: 12px; display: inline-block;
}
.action-unit {
    font-size: 1.8rem; font-weight: 900; color: #111827;
    margin: 12px 0 8px 0; letter-spacing: -0.02em;
}

/* メトリックカード (Glassmorphism) */
.kpi-grid { display:grid; grid-template-columns: repeat(4, 1fr); gap:20px; margin-top:24px; }
.metric-card {
    background: rgba(255, 255, 255, 0.7);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255, 255, 255, 0.5);
    border-radius: 16px;
    padding: 20px;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    transition: transform 0.2s;
}
.metric-card:hover { transform: translateY(-2px); }
.metric-label { color: var(--neutral); font-size: 0.85rem; font-weight: 600; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.05em; }
.metric-value { font-size: 2.2rem; font-weight: 900; color: #0f172a; line-height: 1; }
.metric-sub { font-size: 0.8rem; color: var(--neutral); margin-top: 8px; font-weight: 500; }

/* チャートコンテナ */
.chart-container {
    margin-top: 24px;
}

/* バッジ */
.badge-container { display: flex; align-items: center; gap: 8px; }
.badge {
    background: linear-gradient(135deg, #fef3c7 0%, #fffbeb 100%);
    border: 1px solid #f59e0b;
    color: #b45309;
    padding: 4px 12px;
    border-radius: 99px;
    font-size: 0.8rem;
    font-weight: 700;
    box-shadow: 0 2px 4px rgba(245, 158, 11, 0.1);
    display: inline-flex; align-items: center;
    white-space: nowrap;
}

/* チャートタイトル */
.chart-title {
    font-size: 1.1rem;
    font-weight: 700;
    color: #111827;
    margin-bottom: 12px;
}


/* ============================================
   ライトモード用 SAC/Ant Design タブスタイル
   ============================================ */

/* SAC/Ant Designタブバーの明示的なライトモードスタイル */
.stApp .stTabs,
.stApp .ant-tabs,
.stApp .ant-tabs-top {
    background-color: transparent !important;
}

.stApp .ant-tabs-nav,
.stApp .ant-tabs-nav-wrap,
.stApp .ant-tabs-nav-list {
    background-color: #ffffff;
    background: #ffffff;
    border:1px solid #e5e7eb;
    border-radius: 8px;
    padding: 4px;
}

.stApp .ant-tabs-tab,
.stApp button[data-baseweb="tab"] {
    background-color: transparent !important;
    color: #6b7280 !important;
    border-radius: 6px !important;
}

.stApp .ant-tabs-tab-active,
.stApp button[data-baseweb="tab"][aria-selected="true"] {
    background-color: #3b82f6 !important;
    color: white !important;
}

.stApp .ant-tabs-tab-btn {
    color: inherit !important;
}

/* レスポンシブ */
@media (max-width: 900px) {
  .kpi-grid { grid-template-columns: repeat(2, 1fr); }
  .container { padding: 1rem; }
  .header .container > div { flex-direction: column; align-items: flex-start; gap: 12px; }
  .badge-container { flex-wrap: wrap; }
}

@media (max-width: 768px) {
    /* KPIグリッドを1列に */
    .kpi-grid { grid-template-columns: 1fr; gap: 12px; }
    
    /* 週間プランの横スクロールコンテナ */
    .weekly-plan-container {
        display: flex;
        flex-wrap: nowrap;
        overflow-x: auto;
        gap: 12px;
        padding-bottom: 12px;
        -webkit-overflow-scrolling: touch; /* iOS用スムーズスクロール */
    }
    
    /* 週間プランの各カラム（Streamlitのcolumnはdiv[data-testid="column"]） */
    .weekly-plan-container > div {
        min-width: 140px; /* スマホでの最小幅 */
        flex: 0 0 auto; /* 縮小しない */
    }
    
    /* 暗記カード */
    .flashcard {
        padding: 24px 16px !important;
        min-height: 180px !important;
    }
    .fc-content { font-size: 1.2rem !important; }
    
    /* タブ */
    .stApp .ant-tabs-nav-list {
        display: flex;
        overflow-x: auto;
        white-space: nowrap;
    }
    
    /* 全体の余白調整 */
    .block-container {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }
}
}

/* ============================================
   サイドバーのフォーム要素 - 統一デザインシステム
   ============================================ */

/* サイドバー全体の背景 */
[data-testid="stSidebar"] {
    background-color: #f8fafc;
}

/* 統一されたフォーム要素スタイル */
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stTextInput > div > div > input,
[data-testid="stSidebar"] .stNumberInput > div > div > input,
[data-testid="stSidebar"] .stDateInput > div > div > input,
[data-testid="stSidebar"] .stTextArea > div > div > textarea,
[data-testid="stSidebar"] [data-baseweb="select"] > div,
[data-testid="stSidebar"] [data-baseweb="popover"] {
    background-color: #ffffff !important;
    border: 1.5px solid #94a3b8 !important;
    border-radius: 6px !important;
    color: #0f172a !important;
    font-weight: 500 !important;
    box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05) !important;
}

/* ホバー状態 - 統一 */
[data-testid="stSidebar"] .stSelectbox > div > div:hover,
[data-testid="stSidebar"] .stTextInput > div > div > input:hover,
[data-testid="stSidebar"] .stNumberInput > div > div > input:hover,
[data-testid="stSidebar"] .stDateInput > div > div > input:hover,
[data-testid="stSidebar"] .stTextArea > div > div > textarea:hover {
    border-color: #3b82f6 !important;
    box-shadow: 0 1px 3px 0 rgba(59, 130, 246, 0.1) !important;
}

/* フォーカス状態 - 統一 */
[data-testid="stSidebar"] .stSelectbox > div > div:focus-within,
[data-testid="stSidebar"] .stTextInput > div > div > input:focus,
[data-testid="stSidebar"] .stNumberInput > div > div > input:focus,
[data-testid="stSidebar"] .stDateInput > div > div > input:focus,
[data-testid="stSidebar"] .stTextArea > div > div > textarea:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.12) !important;
    outline: none !important;
}

/* ドロップダウンオプション - 統一 */
[data-testid="stSidebar"] [role="listbox"],
[data-testid="stSidebar"] [role="option"] {
    background-color: #ffffff !important;
    color: #0f172a !important;
}

[data-testid="stSidebar"] [role="option"]:hover {
    background-color: #eff6ff !important;
}

/* セレクトボックスの内部テキスト - 統一 */
[data-testid="stSidebar"] [data-baseweb="select"] span {
    color: #0f172a !important;
    font-weight: 500 !important;
}

/* ラベル - 統一されたコントラスト */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stDateInput label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stTextArea label {
    color: #1e293b !important;
    font-weight: 600 !important;
    font-size: 0.875rem !important;
    margin-bottom: 6px !important;
    display: block !important;
}

/* キャプション - 統一 */
[data-testid="stSidebar"] .stCaption,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
    color: #475569 !important;
    font-weight: 500 !important;
    font-size: 0.875rem !important;
}

/* プレースホルダー - 統一 */
[data-testid="stSidebar"] input::placeholder,
[data-testid="stSidebar"] textarea::placeholder {
    color: #94a3b8 !important;
    opacity: 1 !important;
}

/* ボタン - 統一 */
[data-testid="stSidebar"] button[kind="primary"],
[data-testid="stSidebar"] button[kind="secondary"] {
    border-radius: 6px !important;
    font-weight: 600 !important;
}

/* カスタムラベル - 統一 */
.input-label {
    color: #1e293b;
    font-weight: 600;
    margin-bottom: 6px;
    font-size: 0.875rem;
    display: block;
}
"""

# CSSの組み立て
mode = st.session_state.get("display_mode", "システム設定")
current_theme_name = st.session_state.get("theme", "Blue")
final_css = base_css

# テーマがDarkの場合、またはモードがダークモードの場合
if current_theme_name == "Dark" or mode == "ダークモード":
    final_css += f"\n{dark_css}\n"
# システム設定の場合は、テーマがDarkの場合のみダークモードCSSを適用（Blueなどはライトモード固定）
elif mode == "システム設定" and current_theme_name == "Dark":
    final_css += f"\n{dark_css}\n"

final_css += "</style>"
final_css = final_css.replace("{PRIMARY}", PRIMARY).replace("{ACCENT}", ACCENT).replace("{SUCCESS}", SUCCESS).replace("{WARNING}", WARNING).replace("{DANGER}", DANGER).replace("{BACKGROUND}", BACKGROUND)

st.markdown(final_css, unsafe_allow_html=True)

# ===== ユーザー別ファイルパス定義 =====
user_log_path = f"{st.session_state.user_data_dir}/{st.session_state.current_user}.csv"
user_notes_path = f"{st.session_state.user_data_dir}/{st.session_state.current_user}_notes.csv"

# ===== サイドバー =====
# ===== サイドバー =====
st.sidebar.markdown(f'<div class="chart-header"><i class="bi bi-sliders icon-badge"></i>{t("settings_title")}</div>', unsafe_allow_html=True)

# 1. 企業・目標設定
expanded_settings = not bool(st.session_state.company_name)
with st.sidebar.expander(t("company_goal_settings"), expanded=expanded_settings):
    def save_settings():
        """設定を保存するコールバック"""
        settings = {
            "company_name": st.session_state.comp_input,
            "target_rate_user": st.session_state.target_slider,
            "daily_study_time": st.session_state.time_slider,
            "time_policy": st.session_state.time_select,
            "exam_date": st.session_state.sidebar_exam_date
        }
        # セッションステートも更新
        st.session_state.company_name = settings["company_name"]
        st.session_state.target_rate_user = settings["target_rate_user"]
        st.session_state.daily_study_time = settings["daily_study_time"]
        st.session_state.time_policy = settings["time_policy"]
        st.session_state.exam_date = settings["exam_date"]
        
        # Google Sheetsに保存
        if st.session_state.current_user:
            st.session_state.sheets_manager.save_user_settings(st.session_state.current_user, settings)

    company = st.text_input(t("target_company"), value=st.session_state.company_name, placeholder=t("target_company_placeholder"), key="comp_input", on_change=save_settings)
    
    target = st.slider(t("target_accuracy"), 0, 100, st.session_state.target_rate_user, 5, key="target_slider", on_change=save_settings)
    
    study_time = st.slider(t("daily_study_time"), 10, 180, st.session_state.daily_study_time, 10, key="time_slider", on_change=save_settings)
    
    time_policy = st.selectbox(t("time_policy"), ["標準", "厳しく(-10%)", "緩く(+10%)"], 
                             index=["標準", "厳しく(-10%)", "緩く(+10%)"].index(st.session_state.time_policy), format_func=t, key="time_select", on_change=save_settings)

    # 試験日設定（サイドバーに追加）
    st.markdown("---")
    st.caption(t("exam_date_caption"))
    current_exam_date = st.session_state.exam_date if st.session_state.exam_date else datetime.today()
    new_exam_date = st.date_input(t("exam_date"), value=current_exam_date, key="sidebar_exam_date", on_change=save_settings)

time_factor = {"標準": 1.0, "厳しく(-10%)": 0.9, "緩く(+10%)": 1.1}[st.session_state.time_policy]

# マスタデータの準備（ファイル管理より先にデフォルト読み込み）
if "df_master" not in st.session_state:
    st.session_state.df_master = df_master_default.copy()

# 2. 学習データ入力
expanded_flag = st.session_state.get("expander_open", st.session_state.get("keep_input_open", True))
with st.sidebar.expander(t("input_data_title"), expanded=expanded_flag):
    st.markdown(f"<p class='input-label'>{t('date')}</p>", unsafe_allow_html=True)
    input_date = st.date_input(t("date"), datetime.today(), label_visibility="collapsed", key="dt_input")
    
    # マスタデータ使用
    df_master_use = st.session_state.df_master
    
    subjs = sorted(df_master_use["科目"].unique().tolist())
    # セッションステートからインデックスを復元
    subj_idx = subjs.index(st.session_state.subj) if st.session_state.subj in subjs else 0
    sel_subj = st.selectbox(t("subject"), subjs, index=subj_idx, label_visibility="collapsed", key="s1", format_func=dt)
    
    # 科目変更時のみリセット
    if st.session_state.subj != sel_subj:
        st.session_state.subj = sel_subj
        st.session_state.gen = None
        st.session_state.uni = None
        # trigger_rerun() # 即時反映のため（必要なら）
    
    gens = ["選択"] + sorted(df_master_use[df_master_use["科目"] == st.session_state.subj]["ジャンル"].unique().tolist())
    gen_idx = gens.index(st.session_state.gen) if st.session_state.gen in gens else 0
    sel_gen = st.selectbox(t("genre"), gens, index=gen_idx, label_visibility="collapsed", key="g1", format_func=lambda x: t("select") if x == "選択" else dt(x))
    
    if st.session_state.gen != sel_gen:
        st.session_state.gen = sel_gen
        st.session_state.uni = None
    
    if st.session_state.gen and st.session_state.gen != "選択":
        unis = sorted(df_master_use[(df_master_use["科目"] == st.session_state.subj) & 
                                (df_master_use["ジャンル"] == st.session_state.gen)]["単元"].unique().tolist())
    else:
        unis = []
    
    unis = ["選択"] + unis
    uni_idx = unis.index(st.session_state.uni) if st.session_state.uni in unis else 0
    sel_uni = st.selectbox(t("unit"), unis, index=uni_idx, label_visibility="collapsed", key="u1", format_func=lambda x: t("select") if x == "選択" else dt(x))
    
    if st.session_state.uni != sel_uni:
        st.session_state.uni = sel_uni
    
    ids = df_master_use[(df_master_use["科目"] == st.session_state.subj) & 
                    (df_master_use["ジャンル"] == st.session_state.gen) & 
                    (df_master_use["単元"] == st.session_state.uni)]["問題ID"].tolist() if (
                    st.session_state.uni and st.session_state.uni != "選択") else []
    
    pid = ids[0] if ids else ""
    st.caption(f"{t('problem_id')}: **{pid or t('not_selected')}**")
    
    # タイマー機能
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        def start_timer():
            st.session_state.timer_start_time = time.time()
            st.toast(t("timer_toast_start"), icon="⏱️")

        st.button(t("timer_start"), use_container_width=True, on_click=start_timer)

    with col_t2:
        def stop_timer():
            if st.session_state.get("timer_start_time"):
                elapsed = int(time.time() - st.session_state.timer_start_time)
                st.session_state.timer_elapsed = elapsed
                st.session_state.timer_start_time = None
                st.toast(t("timer_toast_stop").format(elapsed), icon="✅")
            else:
                st.toast(t("timer_toast_warn"), icon="⚠️")

        st.button(t("timer_stop"), use_container_width=True, on_click=stop_timer)
    
    col1, col2 = st.columns(2)
    with col1:
        # タイマー結果があればそれをデフォルトに
        def_at = st.session_state.get("timer_elapsed", 60)
        at = st.number_input(t("answer_time"), min_value=0, max_value=600, value=def_at, step=5, key="at_input")
    with col2:
        cor = st.selectbox(t("result"), ["〇", "✕"], format_func=t, key="cor_select")
    
    cau = st.selectbox(t("miss_reason"), ["-", "理解不足", "知識不足", "時間不足", "ケアレス"], format_func=t, key="cau_select")
    stm = st.number_input(t("study_time_min"), min_value=0, max_value=180, value=10, step=5, key="stm_input")
    
    # 復習メモ欄
    memo = st.text_area(t("memo"), placeholder=t("memo_placeholder"), height=80, key="memo_input")
    
    def add_data_callback(current_pid):
        if not current_pid:
            st.toast(t("toast_error_id"), icon="⚠️")
            return
        
        # 入力値の取得
        input_dt = st.session_state.dt_input
        input_cor = st.session_state.cor_select
        input_at = st.session_state.at_input
        input_cau = st.session_state.cau_select
        input_stm = st.session_state.stm_input
        input_memo = st.session_state.memo_input
        
        new_entry = {
            "日付": input_dt.strftime("%Y-%m-%d"),
            "問題ID": current_pid,
            "正誤": input_cor,
            "解答時間(秒)": input_at,
            "ミスの原因": input_cau,
            "学習投入時間(分)": input_stm
        }
        
        # ログデータ（CSV）への保存
        # 既存のCSVがある場合は読み込んで追記、なければ新規作成
        # スプレッドシートに追加
        success, err = st.session_state.sheets_manager.add_data(st.session_state.current_user, new_entry)
        
        if success:
            # キャッシュをクリアして再読み込みさせるためにリラン
            load_sheet_data.clear()
            
            # ノートがある場合
            if input_memo and input_memo.strip():
                note_entry = {
                    "問題ID": current_pid,
                    "メモ": input_memo.strip(),
                    "登録日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                st.session_state.sheets_manager.add_note(st.session_state.current_user, note_entry)
                load_note_data.clear()
                
            st.session_state.show_success_toast = True
            st.session_state.expander_open = True
        else:
            st.error(f"保存に失敗しました: {err}")

    st.button(t("add_data_btn"), type="primary", use_container_width=True, key="add_btn", on_click=add_data_callback, args=(pid,))

# スプレッドシートマネージャーの初期化
if "sheets_manager" not in st.session_state:
    st.session_state.sheets_manager = GoogleSheetsManager()

# 3. ユーザー管理
with st.sidebar.expander(t("user_management"), expanded=False):
    st.write(f"Logged in as: {st.session_state.current_user}")
    if st.button("Logout"):
        st.session_state.current_user = None
        if os.path.exists('token.json'):
            os.remove('token.json')
        st.rerun()

# 4. ファイルアップロードセクション (CSVインポート機能として残す)
st.sidebar.markdown(f'<div class="chart-header" style="font-size:0.9rem; margin-bottom:8px;"><i class="bi bi-folder icon-badge" style="width:24px; height:24px; font-size:0.9rem;"></i>{t("file_management")}</div>', unsafe_allow_html=True)
with st.sidebar.expander(t("file_details"), expanded=False):
    st.markdown(f"<p class='input-label'>{t('master_csv')}</p>", unsafe_allow_html=True)
    master_file = st.file_uploader(t("master_csv"), type=["csv"], key="master", label_visibility="collapsed")
    
    st.markdown(f"<p class='input-label'>{t('log_csv')}</p>", unsafe_allow_html=True)
    log_file = st.file_uploader(t("log_csv"), type=["csv"], key="log", label_visibility="collapsed")

# マスタ読み込み処理
if master_file:
    try:
        st.session_state.df_master = pd.read_csv(master_file)
        with st.sidebar:
            sac.alert(t("master_loaded"), icon='check-circle', color='success', size='sm')
    except:
        with st.sidebar:
            sac.alert(t("master_failed"), icon='x-circle', color='error', size='sm')

# ログデータの取得（スプレッドシートから）
@st.cache_data(ttl=60)
def load_sheet_data(username):
    return st.session_state.sheets_manager.load_data(username)

df_log_result, load_err = load_sheet_data(st.session_state.current_user)

if load_err:
    df_log = pd.DataFrame(columns=["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"])
    if "シート取得エラー" not in str(load_err):
        st.sidebar.error(f"データ読み込みエラー: {load_err}")
else:
    df_log = df_log_result
    # 必須カラムの存在確認と補完
    required_columns = ["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"]
    for col in required_columns:
        if col not in df_log.columns:
            df_log[col] = pd.Series(dtype='object')

# CSVアップロード時の同期処理
if log_file:
    file_id = f"{log_file.name}_{log_file.size}"
    if st.session_state.get("processed_log_file") != file_id:
        try:
            # 一時ファイルに保存して同期
            with open("temp_upload.csv", "wb") as f:
                f.write(log_file.getbuffer())
            
            success, err = st.session_state.sheets_manager.sync_from_csv(st.session_state.current_user, "temp_upload.csv")
            if success:
                st.session_state.processed_log_file = file_id
                load_sheet_data.clear() # キャッシュクリア
                trigger_rerun()
            else:
                st.sidebar.error(f"同期エラー: {err}")
            
            if os.path.exists("temp_upload.csv"):
                os.remove("temp_upload.csv")
        except Exception as e:
            st.sidebar.error(f"アップロード処理エラー: {str(e)}")

# マニュアル入力用DFも同期
if "df_log_manual" not in st.session_state or st.session_state.get("last_user") != st.session_state.current_user:
    st.session_state.df_log_manual = df_log.copy()
    st.session_state.last_user = st.session_state.current_user

# ノートデータの取得
@st.cache_data(ttl=60)
def load_note_data(username):
    return st.session_state.sheets_manager.load_notes(username)

df_notes_result, note_err = load_note_data(st.session_state.current_user)
if note_err:
    st.session_state.df_notes = pd.DataFrame(columns=["問題ID", "メモ", "登録日時"])
else:
    st.session_state.df_notes = df_notes_result

# マスタデータ変数をローカル変数にセット（後続処理用）
df_master = st.session_state.df_master

# ===== メインコンテンツ =====
# ===== メインコンテンツ =====
badges_html = ""
df = pd.DataFrame()
df_all = pd.DataFrame()

# 変数初期化 (AIコーチなどで使用するため)
att = 0
cor_r = 0.0
tgt_r = st.session_state.target_rate_user / 100
te = 0.0
streak = 0
prediction_text = t("data_insufficient")
prediction_color = "#6B7280"
prediction_sub = t("keep_studying")
bd = pd.DataFrame(columns=["日", "正答率", "ミス", "count", "sum"]) # 初期化

try:
    # データ処理
    df_log["日付"] = pd.to_datetime(df_log["日付"], errors="coerce")
    df_log["解答時間(秒)"] = pd.to_numeric(df_log["解答時間(秒)"], errors="coerce").fillna(0)
    df_log["学習投入時間(分)"] = pd.to_numeric(df_log["学習投入時間(分)"], errors="coerce").fillna(0)
    df_log["ミス"] = (df_log["正誤"] == "✕").astype(int)
    df = pd.merge(df_log, df_master, on="問題ID", how="left")
    df["目標時間"] = df["目標解答時間(秒)"] * time_factor
    
    # カレンダー用（全期間データ）
    df_all = df.copy()

    # 分析期間の選択
    st.sidebar.markdown(f'<div class="chart-header"><i class="bi bi-search icon-badge"></i>{t("analysis_period")}</div>', unsafe_allow_html=True)
    mind = df["日付"].min()
    maxd = df["日付"].max()
    defs = maxd - timedelta(days=7) if pd.notnull(maxd) else datetime.today() - timedelta(days=7)
    sd = st.sidebar.date_input(t("start_date"), defs if pd.notnull(defs) else datetime.today(), key="sd_input")
    ed = st.sidebar.date_input(t("end_date"), maxd if pd.notnull(maxd) else datetime.today(), key="ed_input")

    # 期間フィルタ
    if not df.empty:
        mask = (df["日付"].dt.date >= sd) & (df["日付"].dt.date <= ed)
        df = df.loc[mask]

    # KPI計算
    if not df.empty:
        att = len(df)
        cor_r = 1 - df["ミス"].mean()
        # tgt_r は初期化済み
        
        # 時間超過率
        df["時間超過"] = (df["解答時間(秒)"] > df["目標時間"]).astype(int)
        te = df["時間超過"].mean()
        
        # 集計
        agg = df.groupby("単元").agg({
            "ミス": ["sum", "count"],
            "解答時間(秒)": "mean",
            "目標時間": "mean"
        }).reset_index()
        agg.columns = ["単元", "ミス数", "試行回数", "平均解答時間", "目標時間"]
        agg["正答率"] = (agg["試行回数"] - agg["ミス数"]) / agg["試行回数"]
        
        # 優先度スコア (正答率が低い & 試行回数が多い & 時間がかかる)
        agg["優先度"] = (1 - agg["正答率"]) * 2 + (agg["平均解答時間"] / agg["目標時間"] - 1).clip(0, 1)
        
        # 科目ごとの正答率
        cr = df.groupby("科目")["ミス"].agg(["sum", "count"]).reset_index()
        cr["正答率"] = (cr["count"] - cr["sum"]) / cr["count"]
        
        # 単元ごとの正答率をマージ
        # agg = agg.merge(cr[["単元", "正答率"]], on="単元").sort_values("優先度", ascending=False) # 単元はaggにあるのでマージ不要、科目正答率をどうするか
        # 修正: 科目ごとの正答率は別途表示用。aggは単元別。
        agg = agg.sort_values("優先度", ascending=False)
        
        cs = df.groupby("科目")["ミス"].agg(["sum", "count"]).reset_index()
        cs["正答率"] = (cs["count"] - cs["sum"]) / cs["count"]
        
        bd = df.copy()
        bd["日"] = bd["日付"].dt.date
        bd = bd.groupby("日")["ミス"].agg(["sum", "count"]).reset_index()
        bd["正答率"] = (bd["count"] - bd["sum"]) / bd["count"]
    else:
        agg = pd.DataFrame()
        cs = pd.DataFrame()
        # bd は初期化済み

    # 総演習数（全期間）
    total_att = len(df_all)

    # ===== バッジ判定ロジック =====
    badges = []

    # 1. 初心者 (10問以上)
    if total_att >= 10:
        badges.append(f"<i class='bi bi-egg-fill'></i> {t('beginner_badge')}")

    # 2. 継続日数 (Streak)
    if not df_all.empty:
        dates = sorted(df_all["日付"].dropna().dt.date.unique())
        if len(dates) > 0:
            # streak は初期化済み (0)
            streak = 1
            # 最新の日付が今日か昨日かを確認
            last_d = dates[-1]
            today_d = datetime.today().date()
            
            # もし最新データが昨日より前なら、継続は途切れている（ただし今日はまだやってないだけかもしれないので0にはしないが、連続記録としてはストップ）
            # ここではシンプルに「最新の連続記録」を計算
            
            for i in range(2, len(dates) + 1):
                if (last_d - dates[-i]).days == 1:
                    streak += 1
                    last_d = dates[-i]
                else:
                    break
            
            # 今日か昨日学習していれば継続中とみなす
            if (today_d - dates[-1]).days <= 1:
                badges.append(f"<i class='bi bi-fire'></i> {t('streak_badge').format(streak=streak)}")
            else:
                # 途切れている場合
                badges.append(f"<i class='bi bi-clock-history'></i> {t('last_streak_badge').format(streak=streak)}")

    # 3. 推論マスター (推論ジャンルの正答率80%以上 & 5問以上)
    if not df.empty:
        genre_stats = df.groupby("ジャンル")["ミス"].agg(["sum", "count"])
        genre_stats["acc"] = (genre_stats["count"] - genre_stats["sum"]) / genre_stats["count"]
        for g_name, row in genre_stats.iterrows():
            if row["count"] >= 5 and row["acc"] >= 0.8:
                badges.append(f"<i class='bi bi-trophy-fill'></i> {dt(g_name)}{t('master_suffix')}")

    # 4. スピードスター (平均解答時間が目標の80%以下 & 正答率80%以上)
    if att >= 10 and cor_r >= 0.8:
        avg_time = df["解答時間(秒)"].mean()
        avg_target = df["目標時間"].mean()
        if avg_target > 0 and avg_time <= avg_target * 0.8:
            badges.append(f"<i class='bi bi-lightning-fill'></i> {t('speedster_badge')}")

    # バッジHTML生成（最大3個まで）
    display_badges = badges[:3]  # 最初の3つのみ表示
    for b in display_badges:
        badges_html += f"<span class='badge'>{b}</span>"
    
    # 4つ以上ある場合は「+N」を表示
    if len(badges) > 3:
        remaining = len(badges) - 3
        badges_html += f"<span class='badge' style='background: #e5e7eb; color: #6b7280; border-color: #9ca3af;'>+{remaining}</span>"

except Exception as e:
    st.error(f"{t('data_processing_error')}: {e}")


# ===== ヘッダー (Data Loaded) =====
title_text = t("app_title")
company_val = st.session_state.get('company_name', '')
if not company_val:
    company_val = t("target_company") if st.session_state.language == "English" else t('target_company') # Fallback or just use t()
target_lbl = t("goal_label")
policy_val = st.session_state.get('time_policy',t('standard'))

# カウントダウン
countdown_html = ""
days_left = "-"
if st.session_state.exam_date:
    days_left = (pd.to_datetime(st.session_state.exam_date) - pd.to_datetime(datetime.today().date())).days
    if days_left >= 0:
        lbl = t("days_left")
        unit = t("days_unit")
        bg_col = "#ef4444" if days_left <= 7 else "#3b82f6"
        countdown_html = f"<div style='background:{bg_col}; color:white; padding:2px 10px; border-radius:6px; font-weight:bold; font-size:0.8rem; display:flex; align-items:center; gap:4px;'><span>{lbl}</span><span style='font-size:1rem;'>{days_left}</span><span>{unit}</span></div>"

st.markdown(
    f"<div class='header'><div class='container'>"
    f"<div style='display:flex; justify-content:space-between; align-items:center;'>"
    f"<div class='title-wrap'>"
    f"<div class='logo'><i class='bi bi-journal-text'></i></div>"
    f"<div>"
    f"<div style='display:flex; align-items:center; gap:12px;'>"
    f"<h1 style='color:#1e293b; margin:0; font-size:1.8rem; font-weight:800;'>{title_text}</h1>"
    f"{countdown_html}"
    f"</div>"
    f"<p style='color:#64748b; margin:4px 0 0 0; font-weight:500;'>{company_val} | {target_lbl} {st.session_state.get('target_rate_user',80)}% | {policy_val}</p>"
    f"</div></div>"
    f"<div class='badge-container'>{badges_html}</div>"
    f"</div></div></div>",
    unsafe_allow_html=True
)

if not df.empty:
    cau = df[df["ミス"] == 1]["ミスの原因"].value_counts().reset_index()
    cau.columns = [t("cause"), t("count")]
    cau[t("cause")] = cau[t("cause")].apply(t)

    # --- 3. 合格ライン到達予測 (Linear Regression) ---
    prediction_text = t("data_insufficient")
    prediction_sub = t("min_3_days_data")
    prediction_color = NEUTRAL

    if len(bd) >= 3:
        x = np.arange(len(bd))
        y = bd["正答率"].values
        if np.std(y) == 0:
            prediction_text = t("no_change")
            prediction_sub = t("accuracy_constant")
        else:
            z = np.polyfit(x, y, 1)
            slope = z[0]
            
            if cor_r >= tgt_r:
                prediction_text = t("achieved_exclamation")
                prediction_sub = t("goal_cleared")
                prediction_color = SUCCESS
            elif slope <= 0.001: # ほぼ横ばいか減少
                prediction_text = t("no_improvement")
                prediction_sub = t("review_study_method")
                prediction_color = DANGER
            else:
                intercept = z[1]
                days_needed = (tgt_r - intercept) / slope
                current_day = len(bd) - 1
                days_remaining = days_needed - current_day
                
                if days_remaining <= 0:
                     prediction_text = t("close_to_achieving")
                     prediction_sub = t("almost_there")
                     prediction_color = SUCCESS
                elif days_remaining > 365:
                    prediction_text = t("over_1_year")
                    prediction_sub = t("speed_up_needed")
                    prediction_color = WARNING
                else:
                    pred_date = datetime.today() + timedelta(days=int(days_remaining))
                    prediction_text = pred_date.strftime("%Y/%m/%d")
                    prediction_sub = t("predicted_achievement_date")
                    prediction_color = PRIMARY
    
else:
    cau = pd.DataFrame(columns=[t("cause"), t("count")])
    prediction_text = t("data_insufficient")
    prediction_sub = t("no_data")
    prediction_color = NEUTRAL

# ===== アクション & メニュー (2カラムレイアウト) =====
if not df.empty:
    ac1, ac2 = st.columns(2)

    with ac1:
        tu = agg.iloc[0] if not agg.empty else None
        if tu is not None:
            top_unit_accuracy = tu["正答率"]
            tc = cau.iloc[0][t("cause")] if not cau.empty else t("unknown")
            rsn = f"{t('accuracy_rate')}{top_unit_accuracy:.0%}。" + (t("time_shortage_issue") if te > 0.3 else f"「{tc}」{t('main_cause_review_field')}")
            
            unit_name = tu['単元']
            
            st.markdown(f"""
<div class="action-card" style="height: 100%;">
  <div class="action-icon"><i class="bi bi-lightning-charge-fill"></i></div>
  <div class="action-content">
    <div class="action-header">
      <div class="action-title">{t('next_week_focus_unit')}</div>
      <div class="priority-badge">{t('highest_priority')}</div>
    </div>
    <div class="action-unit">{dt(unit_name)}</div>
    <div class="action-reason">{rsn}</div>
    

  </div>
</div>
""", unsafe_allow_html=True)

    with ac2:
        # 1. 本日の学習メニュー提案
        st.markdown(f"""
        <div class="action-card" style="height: 100%; border-color: {PRIMARY}; background: linear-gradient(135deg, #eff6ff 0%, #ffffff 100%);">
          <div class="action-icon" style="background: {PRIMARY}; box-shadow: 0 4px 6px -1px rgba(59, 130, 246, 0.3);"><i class="bi bi-calendar-event-fill"></i></div>
          <div class="action-content">
            <div class="action-header">
              <div class="action-title">{t('todays_study_menu')}</div>
              <div class="priority-badge" style="background: {PRIMARY};">{t('recommended')}</div>
            </div>
            <div style="margin-top: 12px;">
        """, unsafe_allow_html=True)
        
        if not agg.empty:
            top_3 = agg.head(3)
            for i, row in top_3.iterrows():
                # 優先度に応じて問題数を提案 (例: 優先度1.0 -> 3問, 0.5 -> 2問)
                q_count = max(1, min(5, int(row["優先度"] * 4)))
                st.markdown(f"""
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; border-bottom:1px dashed #e5e7eb; padding-bottom:4px;">
                    <span style="font-weight:700; color:#374151;">{i+1}. {dt(row['単元'])}</span>
                    <span style="font-weight:800; color:{PRIMARY};">{q_count}{t('questions_unit')}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.markdown(f"<div>{t('cannot_propose_no_data')}</div>", unsafe_allow_html=True)
            
        st.markdown("""
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

# ===== KPIカード =====
st.markdown("<div class='container'>", unsafe_allow_html=True)

# モダンなタブナビゲーション (SAC) - テーマに応じて色を変更
current_theme_name = st.session_state.get("theme", "Blue")
is_dark_mode = (current_theme_name == "Dark") or (st.session_state.get("display_mode") == "ダークモード")

# SACタブの背景色を強制的に変更するCSS
if is_dark_mode:
    sac_tab_css = """
    <style>
    /* SAC タブのダークモード強制スタイル - Nuclear (No .stApp dependency) */
    html body div[class*="ant-tabs-nav"],
    html body div[class*="ant-tabs-nav"] * {
        background-color: #1e293b !important;
        background: #1e293b !important;
        border-color: #334155 !important;
    }
    html body div[class*="ant-tabs-tab"] {
        background-color: transparent !important;
        color: #94a3b8 !important;
    }
    html body div[class*="ant-tabs-tab-active"],
    html body div[class*="ant-tabs-tab-active"] * {
        background-color: #3b82f6 !important;
        color: white !important;
    }
    
    /* SAC Divider Fix - Nuclear (No .stApp dependency) */
    html body div[class*="ant-divider"],
    html body div[class*="ant-divider"] * {
        border-top-color: #334155 !important;
        color: #f1f5f9 !important;
    }
    html body div[class*="ant-divider-inner-text"],
    html body div[class*="ant-divider-inner-text"] * {
        background-color: #0f172a !important;
        color: #f1f5f9 !important;
    }
    </style>
    """
    st.markdown(sac_tab_css, unsafe_allow_html=True)

tab_selection = sac.tabs([
    sac.TabsItem(label=t("tab_dashboard"), icon='bar-chart-fill'),
    sac.TabsItem(label=t("tab_data_list"), icon='table'),
    sac.TabsItem(label=t("tab_ai_analysis"), icon='robot'),
    sac.TabsItem(label=t("tab_ai_chat"), icon='chat-dots-fill'),
    sac.TabsItem(label=t("tab_ranking"), icon='trophy-fill'),
    sac.TabsItem(label=t("tab_flashcards"), icon='card-text'),
    sac.TabsItem(label=t("tab_review_notes"), icon='journal-bookmark-fill'),
    sac.TabsItem(label=t("tab_settings"), icon='gear-fill'),
], align='center', size='lg', color='blue')


if tab_selection == t("tab_dashboard"):
    if df_all.empty:
        sac.alert(t("sidebar_input_prompt"), icon='info-circle', color='info')
    else:
        # st.markdown("### 📊 主要指標") # Removed
        
        # AIコーチ (常に表示)
        advice_text = generate_ai_advice(cor_r, tgt_r, te, streak)
        sac.alert(advice_text, icon='lightbulb', color='info', size='sm')

        # ダッシュボード項目のレンダリング関数
        def render_metrics():
            # KPI ストリップ (Unified Design)
            st.markdown("""
            <style>
            .stats-strip {
                background: rgba(255, 255, 255, 0.6);
                backdrop-filter: blur(12px);
                border: 1px solid rgba(255, 255, 255, 0.5);
                border-radius: 16px;
                padding: 20px 0;
                display: flex;
                align-items: center;
                justify-content: space-evenly;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
                margin-top: 16px;
                margin-bottom: 24px;
            }
            .stat-item {
                flex: 1;
                text-align: center;
                border-right: 1px solid rgba(0,0,0,0.06);
                padding: 0 12px; /* 少し詰める */
            }
            .stat-item:last-child { border-right: none; }
            .stat-label { 
                color: var(--neutral); font-size: 0.8rem; font-weight: 600; 
                margin-bottom: 4px; letter-spacing: 0.03em;
            }
            .stat-value { 
                font-size: 1.8rem; font-weight: 900; line-height: 1.1; 
                margin-bottom: 2px;
            }
            .stat-sub { 
                font-size: 0.7rem; color: var(--neutral); font-weight: 500; 
            }
            </style>
            """, unsafe_allow_html=True)

            # 値の計算
            col_cor = SUCCESS if cor_r >= tgt_r else DANGER
            gap = cor_r - tgt_r
            col_gap = SUCCESS if gap >= 0 else DANGER
            col_time = DANGER if te > 0.3 else SUCCESS

            st.markdown(f"""
            <div class="stats-strip">
                <div class="stat-item">
                    <div class="stat-label">{t("current_accuracy")}</div>
                    <div class="stat-value" style="color:{col_cor}">{cor_r:.0%}</div>
                    <div class="stat-sub">{t("period_average")}</div>
                </div>
                <div class="stat-item">
                    <div class="stat-label">{t("gap_to_goal")}</div>
                    <div class="stat-value" style="color:{col_gap}">{gap:+.0%}</div>
                    <div class="stat-sub">{t("achieved") if gap>=0 else t("not_achieved")}</div>
                </div>
                <div class="stat-item">
                    <div class="stat-label">{t("forecast")}</div>
                    <div class="stat-value" style="color:{prediction_color}; font-size: 1.6rem;">{prediction_text}</div>
                    <div class="stat-sub">{prediction_sub}</div>
                </div>
                <div class="stat-item">
                    <div class="stat-label">{t("time_excess_rate")}</div>
                    <div class="stat-value" style="color:{col_time}">{te:.0%}</div>
                    <div class="stat-sub">{t("over_target_time")}</div>
                </div>
                <div class="stat-item">
                    <div class="stat-label">{t("total_exercises")}</div>
                    <div class="stat-value" style="color:var(--primary)">{att}</div>
                    <div class="stat-sub">{t("total_problems")}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        def render_calendar():
            # ===== 学習カレンダー =====
            with st.expander(t("study_calendar"), expanded=True):
                # セッションステートで表示月を管理
                if "calendar_year" not in st.session_state:
                    st.session_state.calendar_year = datetime.now().year
                if "calendar_month" not in st.session_state:
                    st.session_state.calendar_month = datetime.now().month
                
                # 月間ナビゲーション
                c_nav1, c_nav2, c_nav3 = st.columns([1, 5, 1])
                with c_nav1:
                    if st.button(t("prev_month"), key="prev_month"):
                        if st.session_state.calendar_month == 1:
                            st.session_state.calendar_month = 12
                            st.session_state.calendar_year -= 1
                        else:
                            st.session_state.calendar_month -= 1
                        trigger_rerun()
                        
                with c_nav3:
                    if st.button(t("next_month"), key="next_month"):
                        if st.session_state.calendar_month == 12:
                            st.session_state.calendar_month = 1
                            st.session_state.calendar_year += 1
                        else:
                            st.session_state.calendar_month += 1
                        trigger_rerun()
                
                with c_nav2:
                    st.markdown(f"<div style='text-align: center; font-size: 1.1rem; font-weight: 700; padding: 8px;'>{st.session_state.calendar_year}{t('year')}{st.session_state.calendar_month}{t('month')}</div>", unsafe_allow_html=True)
                
                # 週間プランからカレンダー用のデータを生成
                weekly_plan_for_calendar = {}
                if st.session_state.exam_date:
                    weekly_plan_data = generate_weekly_study_plan(
                        df_all, 
                        st.session_state.exam_date, 
                        tgt_r, 
                        cor_r
                    )
                    if weekly_plan_data:
                        weekly_plan_for_calendar = weekly_plan_data
                
                # カレンダー表示
                result = generate_calendar_heatmap(
                    df_all,
                    st.session_state.calendar_year,
                    st.session_state.calendar_month,
                    exam_date=st.session_state.exam_date,
                    weekly_plan=weekly_plan_for_calendar
                )
            
                if result and result[0] and result[1]:
                    css, html = result
                    full_html = css + html
                    import streamlit.components.v1 as components
                    components.html(full_html, height=400, scrolling=False)
                    st.markdown("<div style='margin-top: -80px;'></div>", unsafe_allow_html=True)

        # 設定された順序でウィジェットを表示
        widgets_map = {
            "主要指標": render_metrics,
            "学習カレンダー": render_calendar,
            # "週間学習プラン": render_weekly_plan, # 後で定義
            # "バッジ": render_badges # 後で定義
        }
        
        # しかし、ユーザーは「並び替え」も求めているため、
        # 全てのコンポーネントを関数化してリスト順に呼び出すのがベストです。
        
        # 残りのコンポーネントの関数化（インラインで定義）
        def render_weekly_plan():
            if st.session_state.exam_date:
                sac.divider(label=t('weekly_learning_plan'), icon='calendar-week', align='left')
                col_plan1, col_plan2 = st.columns([3, 1])
                with col_plan1:
                    st.caption(t("weekly_plan_desc"))
                with col_plan2:
                    if st.button(t("update_plan"), key="update_plan_btn"):
                        trigger_rerun()
                
                # Use global df_all if available
                target_df = df_all if 'df_all' in globals() and not df_all.empty else pd.DataFrame()
                
                plan_data = generate_weekly_study_plan(
                    target_df, 
                    st.session_state.exam_date, 
                    st.session_state.target_rate_user / 100, 
                    0
                )
                
                if plan_data:
                    # Pagination Logic
                    plan_items = sorted(plan_data.items(), key=lambda x: x[0])
                    total_days = len(plan_items)
                    DAYS_PER_PAGE = 7
                    
                    # Find today's index
                    today_str = datetime.now().strftime('%Y-%m-%d')
                    today_idx = 0
                    for i, (d_str, _) in enumerate(plan_items):
                        if d_str == today_str:
                            today_idx = i
                            break
                    
                    if "plan_page_idx" not in st.session_state:
                        st.session_state.plan_page_idx = today_idx // DAYS_PER_PAGE
                        
                    start_idx = st.session_state.plan_page_idx * DAYS_PER_PAGE
                    # Boundary check
                    if start_idx >= total_days or start_idx < 0:
                         st.session_state.plan_page_idx = today_idx // DAYS_PER_PAGE
                         start_idx = st.session_state.plan_page_idx * DAYS_PER_PAGE
                    
                    end_idx = min(start_idx + DAYS_PER_PAGE, total_days)
                    
                    
                    # Navigation Buttons
                    c_prev, c_mid, c_next = st.columns([1, 4, 1])
                    with c_prev:
                        if start_idx > 0:
                            if st.button("← " + t("prev_week"), key="plan_prev_btn"):
                                st.session_state.plan_page_idx -= 1
                                trigger_rerun()
                    with c_next:
                        if end_idx < total_days:
                            if st.button(t("next_week") + " →", key="plan_next_btn"):
                                st.session_state.plan_page_idx += 1
                                trigger_rerun()
                    
                    # 週間プラン表示コンテナ（CSSで横スクロール制御）
                    st.markdown('<div class="weekly-plan-container">', unsafe_allow_html=True)
                    
                    current_items = plan_items[start_idx:end_idx]
                    cols = st.columns(len(current_items))

                    # Display Items
                    current_items = plan_items[start_idx:end_idx]
                    cols = st.columns(len(current_items))
                    
                    weekdays = [t("mon"), t("tue"), t("wed"), t("thu"), t("fri"), t("sat"), t("sun")]

                    for i, col in enumerate(cols):
                        date_str, plan = current_items[i]
                        day_date = datetime.strptime(date_str, "%Y-%m-%d")
                        is_today = (date_str == today_str)
                        
                        with col:
                            # Header with increased margin (12px)
                            bg_color = PRIMARY if is_today else "#f3f4f6"
                            text_color = "white" if is_today else "#4b5563"
                            wd = weekdays[day_date.weekday()]
                            
                            st.markdown(f"""
                            <div style="background:{bg_color}; color:{text_color}; padding:4px; border-radius:4px 4px 0 0; text-align:center; font-weight:bold; font-size:0.8rem; width: 94%; margin: 0 auto 12px auto;">
                                {wd}<br><span style="font-size:0.7rem;">{day_date.strftime('%m/%d')}</span>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            # Content
                            units = plan.get('units', [])
                            
                            # カード風コンテナ
                            with st.container():
                                # CSS: Sibling selector approach
                                st.markdown(f"""
                                <style>
                                /* Reduce gap in the vertical block containing plan markers */
                                div[data-testid="stVerticalBlock"]:has(span.plan-marker-lang),
                                div[data-testid="stVerticalBlock"]:has(span.plan-marker-math),
                                div[data-testid="stVerticalBlock"]:has(span.plan-marker-other) {{
                                    gap: 0.25rem !important;
                                }}

                                /* Hide the marker containers so they don't take up space/gaps */
                                div[data-testid="element-container"]:has(span.plan-marker-lang),
                                div[data-testid="element-container"]:has(span.plan-marker-math),
                                div[data-testid="element-container"]:has(span.plan-marker-other) {{
                                    display: none !important;
                                }}

                                /* Language Style (Blue) */
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-lang) + div button {{
                                    background-color: #f0f9ff !important; /* sky-50 */
                                    border: 1px solid #bae6fd !important; /* sky-200 */
                                    border-left: 5px solid #0284c7 !important; /* sky-600 */
                                    color: #0c4a6e !important; /* sky-900 */
                                    border-radius: 6px !important;
                                    padding: 0.25rem 0.5rem !important;
                                    min-height: 3.5rem !important;
                                    height: auto !important;
                                    display: flex !important;
                                    align-items: center !important;
                                    justify-content: flex-start !important;
                                    box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                                    transition: all 0.2s ease;
                                    width: 94% !important;
                                    margin: 0 auto !important;
                                }}
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-lang) + div button:hover {{
                                    background-color: #e0f2fe !important; /* sky-100 */
                                    transform: translateY(-1px);
                                    box-shadow: 0 4px 6px rgba(0,0,0,0.08);
                                }}
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-lang) + div button span[data-testid="stIconMaterial"] {{
                                    color: #0284c7 !important; /* sky-600 */
                                }}

                                /* Non-Language Style (Orange) */
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-math) + div button {{
                                    background-color: #fff7ed !important; /* orange-50 */
                                    border: 1px solid #fed7aa !important; /* orange-200 */
                                    border-left: 5px solid #ea580c !important; /* orange-600 */
                                    color: #7c2d12 !important; /* orange-900 */
                                    border-radius: 6px !important;
                                    padding: 0.25rem 0.5rem !important;
                                    min-height: 3.5rem !important;
                                    height: auto !important;
                                    display: flex !important;
                                    align-items: center !important;
                                    justify-content: flex-start !important;
                                    box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                                    transition: all 0.2s ease;
                                    width: 94% !important;
                                    margin: 0 auto !important;
                                }}
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-math) + div button:hover {{
                                    background-color: #ffedd5 !important; /* orange-100 */
                                    transform: translateY(-1px);
                                    box-shadow: 0 4px 6px rgba(0,0,0,0.08);
                                }}
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-math) + div button span[data-testid="stIconMaterial"] {{
                                    color: #ea580c !important; /* orange-600 */
                                }}

                                /* Other Style (Gray) */
                                div[data-testid="stVerticalBlock"] > div:has(span.plan-marker-other) + div button {{
                                    background-color: #f9fafb !important;
                                    border: 1px solid #e5e7eb !important;
                                    border-left: 5px solid #9ca3af !important;
                                    color: #4b5563 !important;
                                    border-radius: 6px !important;
                                    padding: 0.25rem 0.5rem !important;
                                    min-height: 3.5rem !important;
                                    height: auto !important;
                                    display: flex !important;
                                    align-items: center !important;
                                    justify-content: flex-start !important;
                                    box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                                    width: 94% !important;
                                    margin: 0 auto !important;
                                }}
                                
                                /* Text Wrapping Fix */
                                div[data-testid="stVerticalBlock"] button p {{
                                    white-space: normal !important;
                                    overflow-wrap: break-word !important;
                                    text-align: left !important;
                                    line-height: 1.2 !important;
                                    font-size: 0.8rem !important;
                                    font-weight: 700 !important;
                                    margin: 0 !important;
                                    flex-grow: 1 !important;
                                }}
                                </style>
                                """, unsafe_allow_html=True)
                                
                                for idx, unit in enumerate(units):
                                    unit_name = unit['name']
                                    unit_subj = unit.get('subject', '学習')
                                    unit_type = unit.get('type', '')
                                    
                                    # マーカークラスの決定
                                    if unit_subj in ["言語", "英語"]:
                                        marker_class = "plan-marker-lang"
                                    elif unit_subj in ["非言語", "構造的把握"]:
                                        marker_class = "plan-marker-math"
                                    else:
                                        marker_class = "plan-marker-other"
                                    
                                    # カレンダー追加ポップオーバー（単元名をクリックして開く）
                                    pop_key = f"cal_{date_str}_{idx}"
                                    try:
                                        # マーカーを注入 (非表示)
                                        st.markdown(f'<span class="{marker_class}" style="display:none;"></span>', unsafe_allow_html=True)
                                        
                                        # ラベル: 単元名のみ
                                        btn_label = f"{unit_name}"
                                        
                                        with st.popover(btn_label, icon=":material/event:", use_container_width=True, help=f"{unit_subj}: {t('add_to_google_calendar')}"):
                                            st.markdown(f"**{unit_name}**")
                                            st.caption(f"{t('subject')}: {unit_subj} | {t('type')}: {unit_type}")
                                            
                                            # デフォルト時間は適当に設定（例: 20:00）
                                            sch_time = st.time_input(t("start_time"), value=datetime.strptime("20:00", "%H:%M").time(), key=f"time_{pop_key}")
                                            sch_dur = st.number_input(t("study_duration_min"), value=20, step=10, key=f"dur_{pop_key}")
                                            
                                            if st.button(t("register"), key=f"btn_{pop_key}", type="primary"):
                                                service, error = google_calendar_utils.get_calendar_service()
                                                if error:
                                                    st.error(error)
                                                else:
                                                    try:
                                                        current_year = datetime.now().year
                                                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                                                        
                                                        start_dt = datetime.combine(date_obj, sch_time)
                                                        end_dt = start_dt + timedelta(minutes=sch_dur)
                                                        
                                                        summary = f"📖 {t('study')}: {unit_name}"
                                                        description = f"{t('study_unit')}: {unit_name}\n{t('type')}: {unit_type}"
                                                        
                                                        link, err = google_calendar_utils.add_event_to_calendar(service, summary, start_dt, end_dt, description)
                                                        if link:
                                                            st.success(t("registered_success"))
                                                        elif err:
                                                            st.error(f"{t('error')}: {err}")
                                                    except Exception as e:
                                                        st.error(f"{t('error')}: {e}")
                                        
                                    except Exception:
                                        pass 
                                
                                st.caption(f"Total: {plan['time_minutes']}{t('minutes')}")
                    
                    st.markdown('</div>', unsafe_allow_html=True)

            else:
                st.info(t("set_exam_date_msg"))

        def render_badges():
            st.markdown("---")
            sac.divider(label=t('acquired_badges'), icon='award', align='left')
            
            # バッジ定義
            badge_definitions = [
                {
                    "name": t('beginner_badge'),
                    "icon": "🥚",
                    "desc": "10問以上解答",
                    "condition": lambda df: len(df) >= 10
                },
                {
                    "name": "継続の達人",
                    "icon": "🔥",
                    "desc": "3日以上連続学習",
                    "condition": lambda df: streak >= 3 # streak is calculated globally
                },
                {
                    "name": "推論マスター",
                    "icon": "🏆",
                    "desc": "推論の正答率80%以上",
                    "condition": lambda df: not df[df["ジャンル"]=="推論"].empty and (df[df["ジャンル"]=="推論"]["正誤"]=="〇").mean() >= 0.8
                }
            ]
            
            cols = st.columns(len(badge_definitions))
            for i, badge in enumerate(badge_definitions):
                with cols[i]:
                    is_unlocked = badge["condition"](df_all)
                    opacity = 1.0 if is_unlocked else 0.3
                    grayscale = 0 if is_unlocked else 100
                    
                    st.markdown(f"""
                    <div style="text-align: center; opacity: {opacity}; filter: grayscale({grayscale}%); transition: all 0.3s;">
                        <div style="font-size: 2.5rem; margin-bottom: 8px;">{badge['icon']}</div>
                        <div style="font-weight: 700; font-size: 0.9rem; color: #1f2937; margin-bottom: 4px;">{badge['name']}</div>
                        <div style="font-size: 0.75rem; color: #6b7280; line-height: 1.3;">{badge['desc']}</div>
                    </div>
                    """, unsafe_allow_html=True)

        def render_study_stats():
            # カレンダー下に統計情報を表示
            col1, col2, col3 = st.columns(3)
            
            # 連続学習日数の計算
            if not df_all.empty:
                df_with_date = df_all.copy()
                df_with_date["日付"] = pd.to_datetime(df_with_date["日付"]).dt.date
                unique_dates = sorted(df_with_date["日付"].unique(), reverse=True)
                
                current_streak = 0
                max_streak = 0
                temp_streak = 0
                
                if unique_dates:
                    # 現在の連続日数
                    today = datetime.today().date()
                    if unique_dates[0] == today or (len(unique_dates) > 1 and unique_dates[0] == today - timedelta(days=1)):
                        current_date = unique_dates[0]
                        current_streak = 1
                        for i in range(1, len(unique_dates)):
                            if unique_dates[i] == current_date - timedelta(days=1):
                                current_streak += 1
                                current_date = unique_dates[i]
                            else:
                                break
                    
                    # 最長連続日数
                    for i in range(len(unique_dates)):
                        if i == 0:
                            temp_streak = 1
                        elif unique_dates[i-1] - unique_dates[i] == timedelta(days=1):
                            temp_streak += 1
                        else:
                            max_streak = max(max_streak, temp_streak)
                            temp_streak = 1
                    max_streak = max(max_streak, temp_streak)
                
                # 今月の統計
                today = datetime.today()
                this_month_data = df_with_date[
                    (pd.to_datetime(df_with_date["日付"]).dt.month == today.month) &
                    (pd.to_datetime(df_with_date["日付"]).dt.year == today.year)
                ]
                study_days_this_month = len(this_month_data["日付"].unique())
            else:
                current_streak = 0
                max_streak = 0
                study_days_this_month = 0
            
            # 統計情報をカスタムスタイルで表示
            st.markdown("""
            <style>
            .calendar-stat {
                text-align: center;
                padding: 12px;
                background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
                border-radius: 8px;
                border: 1px solid #e5e7eb;
            }
            .calendar-stat-icon {
                font-size: 1.5rem;
                color: #667eea;
                margin-bottom: 4px;
            }
            .calendar-stat-value {
                font-size: 1.8rem;
                font-weight: 800;
                color: #1f2937;
                margin: 4px 0;
            }
            .calendar-stat-label {
                font-size: 0.8rem;
                color: #6b7280;
                font-weight: 600;
            }
            </style>
            """, unsafe_allow_html=True)
            
            col1_html = f"""
            <div class="calendar-stat">
                <i class="bi bi-fire calendar-stat-icon"></i>
                <div class="calendar-stat-value">{current_streak}{t('days_unit')}</div>
                <div class="calendar-stat-label">{t('current_streak_study')}</div>
            </div>
            """
            
            col2_html = f"""
            <div class="calendar-stat">
                <i class="bi bi-calendar-check calendar-stat-icon"></i>
                <div class="calendar-stat-value">{study_days_this_month}{t('days_unit')}</div>
                <div class="calendar-stat-label">{t('study_days_this_month')}</div>
            </div>
            """
            
            col3_html = f"""
            <div class="calendar-stat">
                <i class="bi bi-trophy calendar-stat-icon"></i>
                <div class="calendar-stat-value">{max_streak}{t('days_unit')}</div>
                <div class="calendar-stat-label">{t('longest_streak_record')}</div>
            </div>
            """
            
            with col1:
                st.markdown(col1_html, unsafe_allow_html=True)
            with col2:
                st.markdown(col2_html, unsafe_allow_html=True)
            with col3:
                st.markdown(col3_html, unsafe_allow_html=True)

        def render_detailed_graphs():
            st.markdown("---")
            st.markdown(f"<div class='chart-header'><i class='bi bi-bar-chart-line-fill icon-badge'></i>{t('widget_detailed_graphs')}</div>", unsafe_allow_html=True)
            
            if df_all.empty:
                st.info(t("no_data_msg"))
                return

            col1, col2 = st.columns(2)
            
            with col1:
                # 1. Subject-wise Proficiency Radar Chart
                st.markdown(f"<div style='margin-bottom:10px; font-weight:bold;'><i class='bi bi-pentagon-half' style='color:#3b82f6;'></i> {t('graph_radar_title')}</div>", unsafe_allow_html=True)
                
                # Calculate accuracy per subject
                df_subj = df_all.copy()
                df_subj["is_correct"] = df_subj["正誤"].apply(lambda x: 1 if x == "〇" else 0)
                subj_acc = df_subj.groupby("科目")["is_correct"].mean().reset_index()
                
                if not subj_acc.empty:
                    categories = subj_acc["科目"].tolist()
                    values = (subj_acc["is_correct"] * 100).tolist()
                    
                    # Close the loop for radar chart
                    categories.append(categories[0])
                    values.append(values[0])
                    
                    fig = go.Figure()
                    fig.add_trace(go.Scatterpolar(
                        r=values,
                        theta=categories,
                        fill='toself',
                        fillcolor='rgba(59, 130, 246, 0.2)',
                        name=t("accuracy_rate"),
                        line=dict(color='#3b82f6', width=3),
                        marker=dict(size=8, color='#3b82f6')
                    ))
                    
                    fig.update_layout(
                        polar=dict(
                            radialaxis=dict(
                                visible=True,
                                range=[0, 100],
                                tickfont=dict(size=10),
                                gridcolor='rgba(0,0,0,0.1)'
                            ),
                            angularaxis=dict(
                                tickfont=dict(size=12, weight="bold")
                            )
                        ),
                        showlegend=False,
                        margin=dict(l=40, r=40, t=20, b=20),
                        height=300,
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)'
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info(t("no_data_msg"))

            with col2:
                # 2. Learning Balance Donut Chart
                st.markdown(f"<div style='margin-bottom:10px; font-weight:bold;'><i class='bi bi-pie-chart-fill' style='color:#8b5cf6;'></i> {t('graph_donut_title')}</div>", unsafe_allow_html=True)
                
                # Count problems per subject
                subj_counts = df_all["科目"].value_counts().reset_index()
                subj_counts.columns = ["科目", "count"]
                total_count = subj_counts["count"].sum()
                
                if not subj_counts.empty:
                    fig2 = px.pie(
                        subj_counts, 
                        values='count', 
                        names='科目', 
                        hole=0.5,
                        color_discrete_sequence=px.colors.qualitative.Set2
                    )
                    fig2.update_traces(
                        textposition='inside', 
                        textinfo='percent+label',
                        marker=dict(line=dict(color='#FFFFFF', width=2))
                    )
                    fig2.update_layout(
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5),
                        margin=dict(l=20, r=20, t=20, b=20),
                        height=300,
                        annotations=[dict(text=f"{total_count}<br>Questions", x=0.5, y=0.5, font_size=20, showarrow=False)]
                    )
                    st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.info(t("no_data_msg"))

        # マッピングの再定義（全関数定義後）
        widgets_map = {
            "主要指標": render_metrics,
            "学習カレンダー": render_calendar,
            "週間学習プラン": render_weekly_plan,
            "バッジ": render_badges,
            "学習記録": render_study_stats
        }

        # 設定された順序でループ実行
        active_widgets = st.session_state.get("dashboard_widgets_v2", ["主要指標", "学習カレンダー", "学習記録", "週間学習プラン"])
             
        for widget_name in active_widgets:
            if widget_name in widgets_map:
                widgets_map[widget_name]()
            else:
                # Handle renamed or removed widgets gracefully
                pass

        # ===== 学習ロードマップ =====
        st.markdown("<div style='margin-top: 24px;'></div>", unsafe_allow_html=True)
        st.markdown(f"<div class='chart-header'><i class='bi bi-signpost-split icon-badge'></i>{t('study_roadmap')}</div>", unsafe_allow_html=True)
        
        roadmap_data, current_phase, recommendations = generate_study_roadmap_detailed(df, st.session_state.df_master)
        
        if roadmap_data and current_phase and recommendations:
            # 現在のフェーズを強調表示
            # キーは日本語（ロジックが返す値）で定義
            phase_colors = {
                "基礎固め": "#3B82F6",
                "標準演習": "#8B5CF6",
                "応用演習": "#EC4899"
            }
            current_color = phase_colors.get(current_phase, "#6B7280")
            
            # 表示用に翻訳
            phase_map = {
                "基礎固め": t("phase_foundation"),
                "標準演習": t("phase_standard"),
                "応用演習": t("phase_advanced")
            }
            display_phase = phase_map.get(current_phase, current_phase)
            
            st.markdown(f"""
            <div style="
                background: linear-gradient(135deg, {current_color}15 0%, {current_color}05 100%);
                border-left: 4px solid {current_color};
                padding: 16px 20px;
                border-radius: 8px;
                margin-bottom: 20px;
            ">
                <div style="font-size: 0.9rem; color: #64748b; font-weight: 600;">{t('current_phase')}</div>
                <div style="font-size: 1.5rem; font-weight: 800; color: {current_color}; margin-top: 4px;">
                    {display_phase}
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # 進捗バーを3つ表示
            col1, col2, col3 = st.columns(3)
            
            for idx, (col, phase_key) in enumerate([(col1, "基礎固め"), (col2, "標準演習"), (col3, "応用演習")]):
                with col:
                    progress = roadmap_data["progress"][idx]
                    accuracy = roadmap_data["accuracy"][idx]
                    status = roadmap_data["status"][idx]
                    
                    # ステータスに応じた色とアイコン（日本語で判定）
                    if status == t("status_completed"):
                        status_color = "#10B981"
                        status_icon = '<i class="bi bi-check-circle-fill" style="color:#10B981;"></i>'
                        status_text_color = "#10B981"
                        display_status = t("completed")
                    elif status == t("status_in_progress"):
                        status_color = "#F59E0B"
                        status_icon = '<i class="bi bi-arrow-repeat" style="color:#F59E0B;"></i>'
                        status_text_color = "#F59E0B"
                        display_status = t("in_progress")
                        display_status = t("in_progress")
                    else:
                        status_color = "#6B7280" # Darker gray for better contrast
                        status_icon = '<i class="bi bi-pause-circle" style="color:#6B7280;"></i>'
                        status_text_color = "#6B7280"
                        display_status = t("not_started")
                    
                        display_status = t("not_started")
                    
                    units_list = "<br>".join([f"・{dt(u)}" for u in roadmap_data["units"][idx]])
                    
                    # フェーズ名の表示用翻訳
                    
                    # フェーズ名の表示用翻訳
                    display_phase_title = phase_map.get(phase_key, phase_key)
                    
                    # Arrow for flow visualization (except last item)
                    arrow_html = ""
                    if idx < 2:
                        arrow_html = f"""
<div style="position: absolute; top: 50%; right: -25px; transform: translateY(-50%); z-index: 10; color: #cbd5e1; font-size: 1.5rem; display: flex; align-items: center; justify-content: center;">
    <i class="bi bi-chevron-right"></i>
</div>
"""
                    
                    st.markdown(f"""
<style>
.roadmap-card {{
    position: relative;
    background: white;
    border: 1px solid {status_color}40;
    border-radius: 12px;
    padding: 16px;
    text-align: center;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
    height: 100%;
    cursor: help;
    transition: all 0.2s ease;
}}
.roadmap-card:hover {{
    transform: translateY(-4px);
    box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
}}
.step-badge {{
    position: absolute;
    top: -10px;
    left: 50%;
    transform: translateX(-50%);
    background: {status_color};
    color: white;
    font-size: 0.75rem;
    font-weight: 700;
    padding: 2px 10px;
    border-radius: 999px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}}
.roadmap-tooltip {{
    visibility: hidden;
    width: 220px;
    background-color: #1e293b;
    color: #fff;
    text-align: left;
    border-radius: 8px;
    padding: 12px;
    position: absolute;
    z-index: 20;
    bottom: 115%;
    left: 50%;
    transform: translateX(-50%);
    opacity: 0;
    transition: opacity 0.2s;
    font-size: 0.8rem;
    line-height: 1.5;
    pointer-events: none;
    box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1);
    border: 1px solid rgba(255,255,255,0.1);
}}
.roadmap-tooltip::after {{
    content: "";
    position: absolute;
    top: 100%;
    left: 50%;
    margin-left: -6px;
    border-width: 6px;
    border-style: solid;
    border-color: #1e293b transparent transparent transparent;
}}
.roadmap-card:hover .roadmap-tooltip {{
    visibility: visible;
    opacity: 1;
}}
</style>

<div style="position: relative; height: 100%;">
    <div class="roadmap-card">
        <div class="step-badge">STEP {idx + 1}</div>
        <div class="roadmap-tooltip">
            <strong style="color: #e2e8f0; display: block; margin-bottom: 4px;">{t('main_units')}</strong>
            {units_list}
        </div>
        <div style="font-size: 2rem; margin-bottom: 12px; margin-top: 8px;">{status_icon}</div>
        <div style="font-weight: 800; font-size: 1.1rem; color: #111827; margin-bottom: 8px;">
            {display_phase_title}
        </div>
        <div style="display: flex; justify-content: space-between; font-size: 0.85rem; color: #334155; margin-bottom: 4px;">
            <span>{t('coverage')}</span>
            <span style="font-weight: 700; color: #0f172a;">{progress:.0f}%</span>
        </div>
        <div style="display: flex; justify-content: space-between; font-size: 0.85rem; color: #334155; margin-bottom: 12px;">
            <span>{t('accuracy_rate')}</span>
            <span style="font-weight: 700; color: #0f172a;">{accuracy:.0f}%</span>
        </div>
        <div style="
            background: #f1f5f9;
            border-radius: 999px;
            height: 8px;
            overflow: hidden;
        ">
            <div style="
                background: {status_color};
                height: 100%;
                width: {accuracy}%;
                border-radius: 999px;
                transition: width 0.5s cubic-bezier(0.4, 0, 0.2, 1);
            "></div>
        </div>
    </div>
    {arrow_html}
</div>
""", unsafe_allow_html=True)
            
            # 次のステップ推薦
            st.markdown("<div style='margin-top: 24px;'></div>", unsafe_allow_html=True)
            st.markdown(f"""
            <div style="
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                padding: 20px;
            ">
                <div style="display:flex; align-items:center; gap:8px; margin-bottom:12px; color:#1e293b; font-weight:700;">
                    <i class="bi bi-lightbulb-fill" style="color:#f59e0b;"></i> {t('next_steps')}
                </div>
                <ul style="margin:0; padding-left:20px; color:#475569;">
                    {''.join([f'<li style="margin-bottom:8px;">{rec}</li>' for rec in recommendations])}
                </ul>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.info(t("roadmap_no_data"))







        # ===== 逆算ロードマップ =====
        if st.session_state.exam_date:
            roadmap_fig = generate_roadmap(st.session_state.exam_date, cor_r, tgt_r)
            if roadmap_fig:
                sac.divider(label=t('roadmap_to_pass'), icon='map', align='center')
                st.plotly_chart(roadmap_fig, use_container_width=True, config={'displayModeBar': False})

        # ===== グラフ =====
        sac.divider(label=t('analysis_graphs'), icon='graph-up', align='center')
        
        m1, m2 = st.columns(2)

        with m1:
            st.markdown(f'<div class="chart-header"><i class="bi bi-graph-up icon-badge"></i>{t("daily_accuracy_trend")}</div>', unsafe_allow_html=True)
            bd = bd.sort_values("日").reset_index(drop=True)
            bd["日_label"] = pd.to_datetime(bd["日"]).dt.day.astype(str) + t("day_suffix")
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=bd["日_label"],
                y=(bd["正答率"] * 100),
                mode='lines+markers+text',
                text=(bd["正答率"] * 100).round(0).astype(int).astype(str) + '%',
                textposition="top center",
                line=dict(color=PRIMARY, width=3, shape='spline'),
                fill='tozeroy',
                fillcolor='rgba(59, 130, 246, 0.1)',
                marker=dict(size=8, color=PRIMARY, line=dict(color='white', width=2)),
                name=t("accuracy_rate"),
                hovertemplate=f'<b>%{{x}}</b><br>{t("accuracy_rate")}：%{{y:.0f}}%<extra></extra>'
            ))
            last_rate = bd["正答率"].iloc[-1] if len(bd) > 0 else cor_r
            target_color = SUCCESS if last_rate >= tgt_r else DANGER
            target_y = tgt_r * 100
            fig.update_layout(shapes=[
                dict(type="line", xref="x", x0=bd["日_label"].iloc[0], x1=bd["日_label"].iloc[-1],
                     yref="y", y0=target_y, y1=target_y,
                     line=dict(color=target_color, width=2, dash="dash"))
            ])
            fig.update_layout(
                template='simple_white',
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                height=300,
                margin=dict(l=40, r=20, t=30, b=40),
                xaxis=dict(showgrid=True, gridcolor='#E6EEF8', tickfont=dict(color='#374151'), zeroline=False),
                yaxis=dict(range=[0, 110], tickmode='array', tickvals=[0, 25, 50, 75, 100],
                           showgrid=True, gridcolor='#E6EEF8', gridwidth=1, zeroline=False),
                hovermode='x unified',
                showlegend=False
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        with m2:
            currentRate_pct = int(round(cor_r * 100))
            targetRate_pct = int(round(tgt_r * 100))
            circumference = 2 * np.pi * 45
            dash = (currentRate_pct / 100.0) * circumference
            remaining = circumference - dash
            svg = f"""
            <div class="metric-card" style="display:flex; align-items:center; justify-content:center; height:300px;">
              <div class="flex flex-col items-center">
                <div class="relative" style="width:160px; height:160px;">
                  <svg viewBox="0 0 100 100" style="transform: rotate(-90deg);">
                    <circle cx="50" cy="50" r="45" fill="none" stroke="var(--border)" stroke-width="8" />
                    <circle cx="50" cy="50" r="45" fill="none" stroke="{SUCCESS if cor_r >= tgt_r else DANGER}" stroke-width="8"
                            stroke-dasharray="{dash:.2f} {remaining:.2f}" stroke-linecap="round" />
                  </svg>
                  <div style="position:absolute; inset:0; display:flex; flex-direction:column; align-items:center; justify-content:center;">
                    <span style="font-size:2rem; font-weight:800; color:var(--card-foreground);">{currentRate_pct}%</span>
                    <span style="font-size:1rem; color:var(--muted-foreground);">/ {targetRate_pct}%</span>
                  </div>
                </div>
              </div>
            </div>
            """
            st.markdown(svg, unsafe_allow_html=True)

        # ===== 下部グラフ =====
        b1, b2 = st.columns(2)

        with b1:
            st.markdown(f'<div class="chart-header"><i class="bi bi-list-check icon-badge"></i>{t("top_5_priority_units")}</div>', unsafe_allow_html=True)
            t5 = agg.head(5).reset_index(drop=True)
            if not t5.empty:
                t5["単元_label"] = t5["単元"].apply(dt)
                max_v = max(t5["優先度"].max(), 1.0)
                pad = max_v * 0.18
                x_max = max_v + pad
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    y=t5["単元_label"],
                    x=[x_max] * len(t5),
                    orientation='h',
                    marker=dict(color='rgba(234,239,243,0.5)'),
                    hoverinfo='none',
                    showlegend=False
                ))
                fig.add_trace(go.Bar(
                    y=t5["単元_label"],
                    x=t5["優先度"],
                    orientation='h',
                    marker=dict(color=PRIMARY, line=dict(color='rgba(0,0,0,0.06)', width=0)),
                    text=t5["優先度"].apply(lambda x: f"{x:.1f}"),
                    textposition='auto',
                    hovertemplate=f'%{{y}}<br>{t("priority")}：%{{x:.2f}}<extra></extra>',
                    name=t('priority')
                ))
                fig.update_layout(
                    template='simple_white',
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    barmode='overlay',
                    height=300,
                    margin=dict(l=100, r=20, t=10, b=20),
                    showlegend=False,
                    xaxis=dict(showgrid=True, gridcolor='rgba(14,30,37,0.06)', range=[0, x_max], zeroline=False),
                    yaxis=dict(autorange='reversed', tickfont=dict(size=14, color='#374151'), dtick=1)
                )
                fig.update_traces(marker_line_width=0)
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        with b2:
            st.markdown(f'<div class="chart-header"><i class="bi bi-pie-chart icon-badge"></i>{t("incorrect_answer_cause_analysis")}</div>', unsafe_allow_html=True)
            fig = go.Figure(go.Bar(
                x=cau[t("cause")],
                y=cau[t("count")],
                text=cau[t("count")],
                textposition='auto',
                marker=dict(color=ACCENT, line=dict(color='rgba(0,0,0,0.06)', width=1)),
                hovertemplate=f'%{{x}}<br>{t("count")}：%{{y}}<extra></extra>'
            ))
            max_y = max(cau[t("count")].max() if not cau.empty else 1, 1)
            fig.update_layout(
                template='simple_white',
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                height=300,
                margin=dict(l=20, r=20, t=10, b=40),
                showlegend=False,
                xaxis=dict(showgrid=False, tickfont=dict(size=12, color='#374151')),
                yaxis=dict(showgrid=True, gridcolor='rgba(14,30,37,0.06)', zeroline=False,
                           tickmode='auto', range=[0, max_y * 1.15], tickfont=dict(size=12, color='#6B7280'))
            )
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

        # ===== AI時系列予測（Prophet） =====
        # Prophet予測（試験日が設定されている場合のみ）
        if st.session_state.get("exam_date") is not None and len(bd) >= 5:
            prophet_result, error_msg = predict_with_prophet(df, tgt_r, st.session_state.exam_date)
            
            if prophet_result:
                st.markdown("---")
                st.markdown(f'<div class="chart-header"><i class="bi bi-graph-up-arrow icon-badge"></i>{t("ai_time_series_prediction_prophet")}</div>', unsafe_allow_html=True)
                st.caption(t("prophet_desc"))
                
                col_p1, col_p2 = st.columns([1, 2])
                
                with col_p1:
                    predicted_rate = prophet_result["predicted_rate"]
                    # 0-1の範囲にクリップ
                    predicted_rate = max(0, min(1, predicted_rate))
                    
                    st.metric(
                        t("exam_day_predicted_accuracy"),
                        f"{predicted_rate:.1%}",
                        delta=f"{(predicted_rate - cor_r):.1%}"
                    )
                    
                    if predicted_rate >= tgt_r:
                        sac.alert(t("goal_achievement_likely"), icon='check-circle', color='success', size='sm')
                    else:
                        gap = tgt_r - predicted_rate
                        sac.alert(f"⚠️ {t('goal_shortage').format(gap=gap)}", icon='exclamation-circle', color='warning', size='sm')
                
                with col_p2:
                    # 予測グラフ（実績 + 予測）
                    forecast_df = prophet_result["forecast"]
                    actual_df = prophet_result["actual_data"]
                    
                    fig_prophet = go.Figure()
                    
                    # 実績データ
                    fig_prophet.add_trace(go.Scatter(
                        x=actual_df["ds"],
                        y=actual_df["y"],
                        mode='markers',
                        name=t('actual_results'),
                        marker=dict(size=8, color=PRIMARY)
                    ))
                    
                    # 予測ライン
                    fig_prophet.add_trace(go.Scatter(
                        x=forecast_df["日付"],
                        y=forecast_df["予測正答率"],
                        mode='lines',
                        name=t('prediction'),
                        line=dict(color=ACCENT, width=2)
                    ))
                    
                    # 信頼区間
                    fig_prophet.add_trace(go.Scatter(
                        x=forecast_df["日付"],
                        y=forecast_df["上限"],
                        mode='lines',
                        name=t('upper_bound'),
                        line=dict(width=0),
                        showlegend=False
                    ))
                    
                    fig_prophet.add_trace(go.Scatter(
                        x=forecast_df["日付"],
                        y=forecast_df["下限"],
                        mode='lines',
                        name=t('lower_bound'),
                        fill='tonexty',
                        fillcolor='rgba(249, 115, 22, 0.2)',
                        line=dict(width=0),
                        showlegend=False
                    ))
                    
                    # 目標ライン
                    fig_prophet.add_hline(
                        y=tgt_r,
                        line_dash="dash",
                        line_color="red",
                        annotation_text=t("goal")
                    )
                    
                    fig_prophet.update_layout(
                        height=250,
                        margin=dict(l=20, r=20, t=20, b=20),
                        yaxis=dict(tickformat=".0%", range=[0, 1.05]),
                        xaxis_title=t("date"),
                        yaxis_title=t("accuracy_rate"),
                        legend=dict(orientation="h", yanchor="top", y=-0.2),
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)'
                    )
                    
                    st.plotly_chart(fig_prophet, use_container_width=True, config={'displayModeBar': False})
            elif error_msg:
                sac.alert(f"{t('prophet_prediction')}: {error_msg}", icon='info-circle', color='info', size='sm')

        # --- 詳細分析（ヒートマップ・散布図） ---
        sac.divider(label=t('detailed_analysis'), icon='search', align='center')
        
        c_h1, c_h2 = st.columns(2)
        with c_h1:
            st.markdown(f'<div class="chart-header"><i class="bi bi-grid-3x3 icon-badge"></i>{t("accuracy_by_field")}</div>', unsafe_allow_html=True)
            heatmap_data = df.groupby(["科目", "ジャンル"])["ミス"].agg(["sum", "count"]).reset_index()
            heatmap_data["正答率"] = (heatmap_data["count"] - heatmap_data["sum"]) / heatmap_data["count"]
            heatmap_matrix = heatmap_data.pivot(index="ジャンル", columns="科目", values="正答率")
            
            # 翻訳適用
            heatmap_matrix.index = [dt(idx) for idx in heatmap_matrix.index]
            heatmap_matrix.columns = [dt(col) for col in heatmap_matrix.columns]
            
            fig_heat = px.imshow(
                heatmap_matrix,
                labels=dict(x=t("subject"), y=t("genre"), color=t("accuracy_rate")),
                x=heatmap_matrix.columns,
                y=heatmap_matrix.index,
                color_continuous_scale="RdBu", # Changed back to RdBu for visibility (Red=Low, Blue=High)
                zmin=0, zmax=1,
                aspect="auto",
                text_auto='.0%' # Show values
            )
            fig_heat.update_traces(xgap=3, ygap=3)
            fig_heat.update_layout(
                template='simple_white',
                height=320, 
                margin=dict(l=0,r=0,t=30,b=0),
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                coloraxis_colorbar=dict(title=t("accuracy_rate"), tickformat=".0%")
            )
            st.plotly_chart(fig_heat, use_container_width=True)
            
        with c_h2:
            st.markdown(f'<div class="chart-header"><i class="bi bi-crosshair icon-badge"></i>{t("weakness_analysis_4_quadrants")}</div>', unsafe_allow_html=True)
            unit_stats = df.groupby("単元").agg({
                "解答時間(秒)": "mean",
                "ミス": ["sum", "count"],
                "科目": "first"
            }).reset_index()
            unit_stats.columns = ["単元", "平均解答時間", "ミス数", "試行回数", "科目"]
            unit_stats["正答率"] = (unit_stats["試行回数"] - unit_stats["ミス数"]) / unit_stats["試行回数"]
            
            # 平均値を計算（象限の基準）
            avg_time = unit_stats["平均解答時間"].mean()
            avg_acc = unit_stats["正答率"].mean()
            max_time = unit_stats["平均解答時間"].max()
            
            fig_scatter = px.scatter(
                unit_stats,
                x="平均解答時間",
                y="正答率",
                size="試行回数",
                color="科目",
                hover_name="単元",
                color_discrete_sequence=[PRIMARY, ACCENT, SUCCESS],
                opacity=0.9
            )
            
            # 象限の背景色（Shapes）
            # 1. 左上 (Ideal): Fast & High Acc
            fig_scatter.add_shape(type="rect", x0=0, y0=avg_acc, x1=avg_time, y1=1.1, fillcolor="rgba(16, 185, 129, 0.1)", layer="below", line_width=0)
            # 2. 右上 (Review): Slow & High Acc
            fig_scatter.add_shape(type="rect", x0=avg_time, y0=avg_acc, x1=max_time*1.2, y1=1.1, fillcolor="rgba(245, 158, 11, 0.1)", layer="below", line_width=0)
            # 3. 左下 (Careless): Fast & Low Acc
            fig_scatter.add_shape(type="rect", x0=0, y0=-0.1, x1=avg_time, y1=avg_acc, fillcolor="rgba(249, 115, 22, 0.1)", layer="below", line_width=0)
            # 4. 右下 (Needs Review): Slow & Low Acc
            fig_scatter.add_shape(type="rect", x0=avg_time, y0=-0.1, x1=max_time*1.2, y1=avg_acc, fillcolor="rgba(239, 68, 68, 0.1)", layer="below", line_width=0)

            # 境界線
            fig_scatter.add_hline(y=avg_acc, line_dash="dash", line_color="#6b7280", opacity=0.5)
            fig_scatter.add_vline(x=avg_time, line_dash="dash", line_color="#6b7280", opacity=0.5)
            
            # 象限ラベル（アノテーション）
            # 左上 (速い・高い): 理想
            fig_scatter.add_annotation(x=avg_time*0.5, y=min(1.0, avg_acc + 0.05), text=t("ideal"), showarrow=False, font=dict(color=SUCCESS, size=12, weight="bold"))
            # 右上 (遅い・高い): 慎重/要反復
            fig_scatter.add_annotation(x=avg_time + (max_time-avg_time)*0.5, y=min(1.0, avg_acc + 0.05), text=t("needs_repetition"), showarrow=False, font=dict(color=WARNING, size=12, weight="bold"))
            # 左下 (速い・低い): ケアレスミス
            fig_scatter.add_annotation(x=avg_time*0.5, y=max(0.0, avg_acc - 0.05), text=t("careless_mistake"), showarrow=False, font=dict(color=ACCENT, size=12, weight="bold"))
            # 右下 (遅い・低い): 基礎不足
            fig_scatter.add_annotation(x=avg_time + (max_time-avg_time)*0.5, y=max(0.0, avg_acc - 0.05), text=t("needs_review"), showarrow=False, font=dict(color=DANGER, size=12, weight="bold"))
            
            fig_scatter.update_traces(marker=dict(line=dict(width=1, color='white')))
            fig_scatter.update_layout(
                template='simple_white',
                height=320, 
                margin=dict(l=0,r=0,t=30,b=0), 
                yaxis=dict(range=[-0.05, 1.05], tickformat=".0%", title=t("accuracy_rate")),
                xaxis=dict(title=t("avg_answer_time_sec"), range=[0, max_time*1.1]),
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            st.plotly_chart(fig_scatter, use_container_width=True)

        # ===== 詳細分析グラフ（科目別習熟度・学習バランス） =====
        render_detailed_graphs()
        
        # ===== 科目別達成状況 =====
        sac.divider(label=t('subject_achievement_status'), icon='stack', align='center')
        if cs.empty:
            st.info(t("no_subject_data"))
        else:
            cols_display = st.columns(len(cs))
            for i, row in enumerate(cs.itertuples()):
                subj_name = row.科目
                with cols_display[i]:
                    key_btn = f"subj_btn_{i}_{subj_name}"
                    clicked = st.button(subj_name, key=key_btn)
                    if clicked:
                        if st.session_state.get("selected_subject") == subj_name:
                            st.session_state.selected_subject = None
                        else:
                            st.session_state.selected_subject = subj_name

                    r = row.正答率
                    delta = r - tgt_r
                    if r >= 1.0:
                        value_col = PRIMARY
                    elif r >= tgt_r:
                        value_col = SUCCESS
                    else:
                        value_col = DANGER
                    delta_col = SUCCESS if delta > 0 else (DANGER if delta < 0 else "#000")
                    width = min(max(int(r * 100), 0), 100)
                    
                    html = f'''
                    <div style="text-align:center; margin-bottom:16px; cursor:pointer;">
                        <div style="font-size:0.9rem; color:#6B7280; margin-bottom:4px;">{subj_name}</div>
                        <div style="font-size:1.5rem; font-weight:700; color:{value_col}; line-height:1;">{r:.0%}</div>
                        <div style="font-size:0.75rem; color:{delta_col}; margin-bottom:8px;">{delta:+.0%}</div>
                        <div style="background-color:#E5E7EB; height:4px; border-radius:2px; width:100%; overflow:hidden;">
                            <div style="background-color:{value_col}; height:100%; width:{width}%;"></div>
                        </div>
                    </div>
                    '''
                    st.markdown(html, unsafe_allow_html=True)

            sel = st.session_state.get("selected_subject", None)
            if sel:
                sac.divider(label=f'<i class="bi bi-search"></i> {sel} {t("unit_accuracy_rate")}', icon='search', align='left')
                units = df[df["科目"] == sel].groupby("単元")["ミス"].agg(["sum", "count"]).reset_index()
                if units.empty:
                    st.info(t("no_data_for_subject"))
                else:
                    units["正答率"] = (units["count"] - units["sum"]) / units["count"]
                    units = units.sort_values("正答率", ascending=False).reset_index(drop=True)
                    
                    # Translate unit names
                    # Keep original for search query if needed, but here we use translated for simplicity or add logic
                    # Actually, for better search results in Japan, maybe we should keep Japanese?
                    # But the user might be English speaker.
                    # Let's use the translated name for now.
                    units["単元"] = units["単元"].apply(dt)
                    
                    # Add search link
                    units["link"] = units["単元"].apply(lambda x: f"https://www.youtube.com/results?search_query={urllib.parse.quote('SPI ' + x)}")
                    units["google_link"] = units["単元"].apply(lambda x: f"https://www.google.com/search?q={urllib.parse.quote('SPI ' + x + t('search_suffix'))}")
                    
                    # Select raw columns and rename for display
                    # HTML Table Generation for Unit Accuracy
                    table_html = f"""
<div style="overflow-x: auto;">
<table style="width:100%; border-collapse: collapse; font-size:0.9rem;">
<thead>
<tr style="border-bottom:2px solid #e5e7eb; color:#6b7280; font-size:0.85rem;">
<th style="padding:12px 8px; text-align:left;">{t("unit")}</th>
<th style="padding:12px 8px; text-align:left; width:40%;">{t("accuracy_rate")}</th>
<th style="padding:12px 8px; text-align:center;">{t("attempts")}</th>
<th style="padding:12px 8px; text-align:center;">{t("resources")}</th>
</tr>
</thead>
<tbody>
"""
                    
                    for _, row in units.iterrows():
                        unit_name = row["単元"]
                        acc = row["正答率"]
                        attempts = row["count"]
                        link = row["link"]
                        google_link = row["google_link"]
                        
                        # Accuracy Bar Color
                        if acc >= 0.8: bar_color = "#10b981" # Success
                        elif acc >= 0.6: bar_color = "#3b82f6" # Primary
                        else: bar_color = "#ef4444" # Danger
                        
                        acc_pct = acc * 100
                        
                        table_html += f"""
<tr style="border-bottom:1px solid #f3f4f6;">
<td style="padding:12px 8px; font-weight:600; color:#1f2937;">{unit_name}</td>
<td style="padding:12px 8px;">
<div style="display:flex; align-items:center; gap:12px;">
<div style="flex-grow:1; background:#f3f4f6; height:8px; border-radius:4px; overflow:hidden;">
<div style="width:{acc_pct}%; background:{bar_color}; height:100%;"></div>
</div>
<span style="font-weight:700; color:#374151; min-width:40px; text-align:right;">{acc_pct:.0f}%</span>
</div>
</td>
<td style="padding:12px 8px; text-align:center; color:#4b5563; font-weight:500;">{attempts}</td>
<td style="padding:12px 8px; text-align:center;">
<a href="{link}" target="_blank" style="text-decoration:none; color:#ef4444; font-size:1.3rem; transition: opacity 0.2s;" onmouseover="this.style.opacity='0.7'" onmouseout="this.style.opacity='1'">
<i class="bi bi-youtube"></i>
</a>
<a href="{google_link}" target="_blank" style="text-decoration:none; color:#3b82f6; font-size:1.2rem; margin-left:12px; transition: opacity 0.2s;" onmouseover="this.style.opacity='0.7'" onmouseout="this.style.opacity='1'">
<i class="bi bi-google"></i>
</a>
</td>
</tr>
"""
                    
                    table_html += "</tbody></table></div>"
                    st.markdown(table_html, unsafe_allow_html=True)

                    # Close button
                    st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
                    if st.button(t("close"), key=f"close_subj_{sel}"):
                        st.session_state.selected_subject = None
                        trigger_rerun()
if tab_selection == t("tab_data_list"):
    st.markdown(f"### 📋 {t('tab_data_list')}")
    sac.divider(label=t('data_download'), icon='download', align='center')
    col_dl1, col_dl2, col_dl3 = st.columns(3)
    
    with col_dl1:
        csv_log = st.session_state.df_log_manual.to_csv(index=False).encode('utf-8-sig')
        st.download_button(t("learning_log_csv"), data=csv_log, file_name=f"study_log_{st.session_state.current_user}.csv", mime="text/csv", use_container_width=True)
        
    with col_dl2:
        if not st.session_state.df_notes.empty:
            csv_notes = st.session_state.df_notes.to_csv(index=False).encode('utf-8-sig')
            st.download_button(t("review_notes_csv"), data=csv_notes, file_name=f"review_notes_{st.session_state.current_user}.csv", mime="text/csv", use_container_width=True)
        else:
            st.button(t("review_notes_none"), disabled=True, use_container_width=True)
            
    with col_dl3:
        if not agg.empty:
            csv_agg = agg.to_csv(index=False).encode('utf-8-sig')
            st.download_button(t("unit_summary_csv"), data=csv_agg, file_name=f"unit_stats_{st.session_state.current_user}.csv", mime="text/csv", use_container_width=True)
        else:
            st.button(t("unit_summary_none"), disabled=True, use_container_width=True)
    
    with st.expander(t("entered_data_list"), expanded=True):
        # 必須カラムの保証
        required_columns = ["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"]
        for col in required_columns:
            if col not in st.session_state.df_log_manual.columns:
                st.session_state.df_log_manual[col] = pd.Series(dtype='object')

        # --- 新機能: データエディタで直接編集 ---
        sac.alert(t("edit_cell_instruction"), icon='pencil-square', color='info', size='sm')
        
        # 日付カラムを datetime に変換してエディタに渡す
        df_editor = st.session_state.df_log_manual.copy()
        if "日付" in df_editor.columns:
            df_editor["日付"] = pd.to_datetime(df_editor["日付"], errors="coerce")

        edited_df = st.data_editor(
            df_editor,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "日付": st.column_config.DateColumn(t("date"), format="YYYY-MM-DD"),
                "正誤": st.column_config.SelectboxColumn(t("result"), options=["〇", "✕"]),
                "ミスの原因": st.column_config.SelectboxColumn(t("miss_reason"), options=["-", "理解不足", "知識不足", "時間不足", "ケアレス"]),
            }
        )
        
        # 編集があった場合、日付を文字列に戻して保存
        if not edited_df.equals(df_editor):
            edited_df["日付"] = edited_df["日付"].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notnull(x) else "")
            st.session_state.df_log_manual = edited_df
            
            # Google Sheetsに保存
            try:
                # 一時ファイルに保存
                edited_df.to_csv("temp_manual_edit.csv", index=False)
                success, err = st.session_state.sheets_manager.sync_from_csv(st.session_state.current_user, "temp_manual_edit.csv")
                
                if success:
                    st.success(t("changes_saved"))
                    load_sheet_data.clear() # キャッシュクリア
                    if os.path.exists("temp_manual_edit.csv"):
                        os.remove("temp_manual_edit.csv")
                    trigger_rerun()
                else:
                    st.error(f"保存エラー: {err}")
            except Exception as e:
                st.error(f"保存処理エラー: {str(e)}")
        
        csv = st.session_state.df_log_manual.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label=t("download_csv"),
            data=csv,
            file_name=f"spi_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
            
    st.markdown("---")
    uploaded = st.file_uploader(t("replace_data_csv"), type=["csv"], key="tab2_upload")
    if uploaded is not None:
        try:
            df_new = pd.read_csv(uploaded)
            required = ["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"]
            missing = [c for c in required if c not in df_new.columns]
            if missing:
                st.error(t("missing_csv_columns").format(columns=', '.join(missing)))
            else:

                st.session_state.df_log_manual = df_new[required].copy()
                
                # Google Sheetsに保存
                try:
                    # 一時ファイルに保存
                    st.session_state.df_log_manual.to_csv("temp_upload_manual.csv", index=False)
                    success, err = st.session_state.sheets_manager.sync_from_csv(st.session_state.current_user, "temp_upload_manual.csv")
                    
                    if success:
                        st.success(t("session_data_replaced"))
                        load_sheet_data.clear() # キャッシュクリア
                        if os.path.exists("temp_upload_manual.csv"):
                            os.remove("temp_upload_manual.csv")
                        trigger_rerun()
                    else:
                        st.error(f"保存エラー: {err}")
                except Exception as e:
                    st.error(f"保存処理エラー: {str(e)}")
        except Exception as e:
            st.error(t("csv_read_failed").format(error=e))
    
    st.markdown("---")
    with st.expander(t("dangerous_operations")):
        if st.button(t("delete_all_data"), type="primary"):
            st.session_state.df_log_manual = pd.DataFrame(columns=["日付", "問題ID", "正誤", "解答時間(秒)", "ミスの原因", "学習投入時間(分)"])
            st.success(t("all_logs_deleted"))
            trigger_rerun()
            
        else:
            st.info(t("no_input_data_prompt"))

if tab_selection == t("tab_ai_analysis"):
    sac.divider(label=t('ai_analysis_report'), icon='robot', align='center')
    st.write(t("ai_analysis_desc"))
    
    if df.empty or len(df) < 5:
        sac.alert(t("ai_analysis_min_data"), icon='exclamation-triangle', color='warning')
    else:
        # 詳細インサイト表示（Phase 2: ルールベースAI強化）
        st.markdown ("---")
        sac.divider(label=t('personalized_learning_analysis'), icon='person-check-fill', align='left')
        st.caption(t("personalized_learning_analysis_desc"))
        
        insights = generate_detailed_insights(df, cor_r, tgt_r, st.session_state.get("exam_date"))
        
        if insights:
            # 優先度別に色分け
            priority_colors = {
                "urgent": "error",
                "high": "warning",
                "medium": "info",
                "low": "success"
            }
            
            for insight in insights:
                icon = insight.get("icon", "info-circle")
                priority = insight.get("priority", "medium")
                color = priority_colors.get(priority, "info")
                message = insight.get("message", "")
                
                sac.alert(
                    message,
                    icon=icon,
                    color=color,
                    banner=True if priority == "urgent" else False,
                    closable=False
                )
        else:
            sac.alert(t("data_accumulation_needed"), icon='info-circle', color='info')
        
        st.markdown("---")
        
        # モデル学習
        with st.spinner(t("ai_model_training")):
            model_acc, importances, encoders = train_ai_models(df)
            
        if model_acc is None:
            st.error(t("model_training_failed"))
        else:
            le_subj, le_unit, min_date = encoders
            
            # 1. 未来予測
            sac.divider(label=t('accuracy_prediction_simulation'), icon='graph-up-arrow', align='left')
            col_ai1, col_ai2 = st.columns([1, 2])
            with col_ai1:
                target_date = st.date_input(t("prediction_date"), value=datetime.today() + timedelta(days=7))
                days_future = (pd.to_datetime(target_date) - min_date).days
                
                # 予測用ダミーデータ作成（平均的な学習条件で予測）
                avg_time = df["解答時間(秒)"].mean()
                avg_study = df["学習投入時間(分)"].mean()
                
                X_pred = []
                # 学習データに含まれるユニークな科目・単元のペアを取得
                unique_pairs = df[["科目", "単元"]].drop_duplicates()
                
                for _, row in unique_pairs.iterrows():
                    try:
                        s_c = le_subj.transform([row["科目"]])[0]
                        u_c = le_unit.transform([row["単元"]])[0]
                        X_pred.append([days_future, s_c, u_c, avg_time, avg_study])
                    except:
                        pass
                
                if X_pred:
                    pred_accs = model_acc.predict(X_pred)
                    final_pred = np.mean(pred_accs)
                    
                    st.metric(t("predicted_accuracy"), f"{final_pred:.1%}", delta=f"{(final_pred - cor_r):.1%}")
                    if final_pred >= tgt_r:
                        sac.alert(t("goal_achievement_likely"), icon='check-circle', color='success', size='sm')
                    else:
                        sac.alert(t("goal_not_achieved"), icon='exclamation-circle', color='warning', size='sm')
            
            with col_ai2:
                # 予測推移グラフ（向こう30日）
                future_days = range(days_future, days_future + 30)
                future_preds = []
                for d in future_days:
                    # 各日の予測（全単元平均）
                    X_d = [[d, x[1], x[2], x[3], x[4]] for x in X_pred]
                    preds = model_acc.predict(X_d)
                    future_preds.append(np.mean(preds))
                
                fig_pred = px.line(x=[min_date + timedelta(days=d) for d in future_days], y=future_preds, 
                                   labels={"x": t("date"), "y": t("predicted_accuracy")}, title=t("30_day_growth_prediction"))
                fig_pred.add_hline(y=tgt_r, line_dash="dash", line_color="red", annotation_text=t("goal"))
                fig_pred.update_layout(height=250, margin=dict(l=20, r=20, t=30, b=20))
                st.plotly_chart(fig_pred, use_container_width=True)

            # 2. 要因分析
            sac.divider(label=t('performance_factor_analysis'), icon='bar-chart-steps', align='left')
            st.caption(t("performance_factor_analysis_desc"))
            fig_imp = px.bar(importances, x="importance", y="feature", orientation="h", 
                             title=t("impact_on_accuracy"), labels={"importance": t("importance"), "feature": t("factor")})
            fig_imp.update_layout(height=200, margin=dict(l=20, r=20, t=30, b=20))
            st.plotly_chart(fig_imp, use_container_width=True)
            
            # 3. AIレコメンド
            sac.divider(label=t('recommended_curriculum'), icon='journal-check', align='left')
            st.caption(t("recommended_curriculum_desc"))
            
            # 全単元の現在の予測正答率を計算
            current_days = (datetime.today() - min_date).days
            recs = []
            for _, row in unique_pairs.iterrows():
                try:
                    s_c = le_subj.transform([row["科目"]])[0]
                    u_c = le_unit.transform([row["単元"]])[0]
                    # 今日の予測
                    p = model_acc.predict([[current_days, s_c, u_c, avg_time, avg_study]])[0]
                    recs.append({t("subject"): dt(row["科目"]), t("unit"): dt(row["単元"]), t("predicted_accuracy"): p})
                except:
                    pass
            
            df_recs = pd.DataFrame(recs)
            # 成長ゾーン (40% - 75%)
            df_growth = df_recs[(df_recs[t("predicted_accuracy")] >= 0.4) & (df_recs[t("predicted_accuracy")] <= 0.75)].sort_values(t("predicted_accuracy"))
            
            if not df_growth.empty:
                for i, row in df_growth.head(3).iterrows():
                    sac.alert(f"**{row[t('subject')]} - {row[t('unit')]}** ({t('predicted_accuracy')}: {row[t('predicted_accuracy')]:.1%})", icon='fire', color='info')
            else:
                sac.alert(t("no_growth_zone_units"), icon='check2-circle', color='success')
            
            # --- 類似問題生成 ---
            sac.divider(label=t("ai_problem_gen_title"), icon='pencil-fill', align='left')
            st.caption(t("ai_problem_gen_desc"))
            
            col_gen1, col_gen2 = st.columns([2, 1])
            with col_gen1:
                # 全単元を取得（マスタデータから）
                all_units = sorted(df_master["単元"].unique().tolist()) if not df_master.empty else ["推論", "集合", "確率"]
                
                # 苦手な単元（正答率が低い順）を優先的に表示するためのソート
                if not agg.empty:
                    weak_units = agg.sort_values("正答率")["単元"].tolist()
                # 苦手な順 + それ以外の単元
                    sorted_units = weak_units + [u for u in all_units if u not in weak_units]
                else:
                    sorted_units = all_units
                
                target_unit = st.selectbox(t("select_unit_label"), sorted_units, format_func=dt)
            
            with col_gen2:
                st.write("") # Spacer
                st.write("")
                if st.button(t("generate_problem_btn"), type="primary", use_container_width=True):
                    with st.spinner(t("generating_problem_spinner")):
                        problem_text = ai_utils.generate_similar_problem("SPI", target_unit)
                        st.session_state.generated_problem = problem_text
            
            if "generated_problem" in st.session_state:
                st.markdown(t("generated_problem_title"))
                st.info(st.session_state.generated_problem)
            
            # 4. 学習フロー可視化（積み上げ棒グラフ）
            st.markdown("---")
            sac.divider(label=t('learning_flow_visualization'), icon='bar-chart-steps', align='left')
            st.caption(t("learning_flow_visualization_desc"))
            
            bar_fig = generate_stacked_bar_chart(df)
            if bar_fig:
                st.plotly_chart(bar_fig, use_container_width=True, config={'displayModeBar': False})
                
                # インサイト表示
                correct_rate = (df["正誤"] == "〇").sum() / len(df)
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); 
                            padding: 16px; border-radius: 12px; border-left: 4px solid {PRIMARY}; margin-top: 16px;">
                    <div style="font-weight: 600; color: #1f2937; margin-bottom: 8px;">
                        <i class="bi bi-lightbulb-fill" style="color: {PRIMARY}; margin-right: 8px;"></i>
                        {t('flow_analysis_insights')}
                    </div>
                    <div style="color: #374151; font-size: 0.9rem;">
                        • {t('overall_accuracy_rate')}: <strong>{correct_rate:.1%}</strong><br>
                        • {t('thick_flow_explanation')}<br>
                        • {t('green_red_flow_explanation')}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                sac.alert(t("data_insufficient_sankey"), icon='info-circle', color='info')

if tab_selection == t("tab_ai_chat"):
    sac.divider(label=t("ai_coach_title"), icon='robot', align='left', size='lg', color='blue')
    st.caption(t("ai_chat_desc"))
    
    # PDFアップロード
    with st.expander(t("upload_pdf_expander"), icon=":material/upload_file:"):
        # Note: st.expander icon supports emojis or Material Symbols (Streamlit 1.34+). 
        # Bootstrap icons are not supported in st.expander icon argument directly unless using emoji shortcodes that map to icons, which is rare.
        # Streamlit supports Material Symbols like ":material/upload_file:".
        # If user strictly wants Bootstrap icons everywhere, we can't do it in st.expander icon easily.
        # But removing the emoji "📂" is a good start.
        # I will use a Material Symbol which is the modern Streamlit way, or just no icon if preferred.
        # The user asked for "Bootstrap icon", but Streamlit native components don't support BI classes.
        # sac components do.
        # I will use sac.divider for headers.
        # For expander, I will remove the emoji from the label.
        uploaded_file = st.file_uploader(t("select_pdf"), type="pdf")
        if uploaded_file is not None:
            # テキスト抽出（キャッシュしておくと良いが、簡易実装として毎回読むか、session_stateに入れる）
            # ファイルが変わった場合のみ読み込むロジック
            if "current_pdf_name" not in st.session_state or st.session_state.current_pdf_name != uploaded_file.name:
                with st.spinner(t("reading_pdf")):
                    pdf_text = ai_utils.extract_text_from_pdf(uploaded_file)
                    st.session_state.pdf_context = pdf_text
                    st.session_state.current_pdf_name = uploaded_file.name
                st.success(t("pdf_read_success").format(uploaded_file.name))
            
            # 読み込み済みであることを表示
            if "pdf_context" in st.session_state:
                st.info(t("current_pdf").format(st.session_state.current_pdf_name))

    # チャット履歴の初期化
    if "messages" not in st.session_state:
        st.session_state.messages = []

    # 履歴の表示
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # ユーザー入力
    if prompt := st.chat_input(t("chat_placeholder")):
        # ユーザーのメッセージを表示
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        # AIの応答
        with st.chat_message("assistant"):
            with st.spinner(t("thinking")):
                # コンテキスト（学習状況）の作成
                context = ""
                if not df_log.empty:
                    total_time = df_log["学習投入時間(分)"].sum()
                    acc = (df_log["正誤"] == "〇").mean()
                    context = t("context_summary").format(total_time, acc)
                
                # PDFコンテキスト
                doc_content = st.session_state.get("pdf_context", "")
                
                response = ai_utils.get_gemini_response(prompt, context, doc_content)
                st.markdown(response)
        
        # 履歴に追加
        st.session_state.messages.append({"role": "assistant", "content": response})

def render_flashcards():
    """
    暗記カード機能のレンダリング
    """
    sac.divider(label=t("flashcards_title"), icon='card-text', align='left', size='lg', color='indigo')
    st.caption(t("flashcards_desc"))
    
    # 単元選択
    units = list(FLASHCARD_DATA.keys())
    selected_unit = st.selectbox(t("select_unit"), units, key="fc_unit_select", format_func=dt)
    
    if selected_unit:
        cards = FLASHCARD_DATA[selected_unit]
        
        # セッション状態でカードインデックスを管理
        if "fc_index" not in st.session_state:
            st.session_state.fc_index = 0
        if "fc_flipped" not in st.session_state:
            st.session_state.fc_flipped = False
        if "fc_shuffled_cards" not in st.session_state:
            st.session_state.fc_shuffled_cards = cards
            
        # ユニットが変わったらリセット
        if st.session_state.get("fc_current_unit") != selected_unit:
            st.session_state.fc_current_unit = selected_unit
            st.session_state.fc_index = 0
            st.session_state.fc_flipped = False
            st.session_state.fc_shuffled_cards = cards
            
        current_cards = st.session_state.fc_shuffled_cards
        total_cards = len(current_cards)
        current_idx = st.session_state.fc_index
        
        if total_cards == 0:
            st.info(t("no_cards_for_unit"))
            return

        card = current_cards[current_idx]
        
        # Import components for HTML embedding
        import streamlit.components.v1 as components
        
        # カード表示エリア
        # レイアウト変更: カードを上に、ボタンを下に配置
        
        # 1. カード表示
        # Client-side Flashcard with HTML/CSS/JS
        
        # Prepare content
        q_text = card['question']
        a_text = card['answer']
        note_text = card.get('note', '')
        sub_q = t("question")
        sub_a = t("answer")
        hint_text = t("click_to_show_answer")
        
        # Determine initial class based on python state
        initial_class = "flipped" if st.session_state.fc_flipped else ""
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <style>
            body {{
                font-family: "Source Sans Pro", sans-serif;
                background-color: transparent;
                margin: 0;
                padding: 0;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 320px;
                perspective: 1000px;
            }}
            .flashcard-container {{
                width: 100%;
                height: 100%;
                position: relative;
                cursor: pointer;
                transform-style: preserve-3d;
                transition: transform 0.6s;
            }}
            .flashcard-container.flipped {{
                transform: rotateY(180deg);
            }}
            .face {{
                position: absolute;
                width: 100%;
                height: 100%;
                backface-visibility: hidden;
                border-radius: 1.5rem;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                text-align: center;
                padding: 30px;
                box-sizing: border-box;
                box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.1);
                border: 1px solid rgba(0,0,0,0.05);
            }}
            .front {{
                background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
                color: #1e293b;
            }}
            .back {{
                background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
                color: #1e3a8a;
                transform: rotateY(180deg);
                border: 2px solid #3b82f6;
            }}
            .fc-sub {{
                font-size: 0.9rem;
                color: #64748b;
                margin-bottom: 1rem;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.1em;
            }}
            .fc-content {{
                font-size: 2rem;
                font-weight: 800;
                margin-bottom: 1rem;
                line-height: 1.4;
            }}
            .fc-note {{
                font-size: 1rem;
                color: #475569;
                margin-top: 1.5rem;
                background: rgba(255,255,255,0.8);
                padding: 12px 16px;
                border-radius: 8px;
                border: 1px solid #e2e8f0;
            }}
            .hint {{
                font-size: 0.9rem;
                color: #94a3b8;
                margin-top: 1.5rem;
                font-weight: 500;
            }}
        </style>
        </head>
        <body>
            <div class="flashcard-container {initial_class}" onclick="this.classList.toggle('flipped')">
                <div class="face front">
                    <div class="fc-sub">{sub_q}</div>
                    <div class="fc-content">{q_text}</div>
                    <div class="hint">{hint_text}</div>
                </div>
                <div class="face back">
                    <div class="fc-sub">{sub_a}</div>
                    <div class="fc-content">{a_text}</div>
                    <div class="fc-note">{note_text}</div>
                </div>
            </div>
        </body>
        </html>
        """
        
        components.html(html_content, height=330)
        
        # 2. コントロールボタン（カードの下に配置）
        st.write(f"**{t('card_counter').format(current_idx + 1, total_cards)}**")
        
        # Custom CSS for buttons
        st.markdown("""
        <style>
        div.stButton > button {
            width: 100%;
            border-radius: 12px;
            border: 1px solid #e5e7eb;
            background-color: white;
            color: #4b5563;
            font-weight: 600;
            box-shadow: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
            transition: all 0.2s;
        }
        div.stButton > button:hover {
            border-color: #3b82f6;
            color: #3b82f6;
            background-color: #eff6ff;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            transform: translateY(-1px);
        }
        div.stButton > button:active {
            transform: translateY(0);
        }
        </style>
        """, unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        
        with c1:
            # Prev
            label_prev = t('prev_card').replace('⬅', '').replace('➡', '').strip()
            if st.button(label_prev, key=f"fc_prev_{current_idx}", icon=":material/arrow_back:", use_container_width=True):
                st.session_state.fc_index = (current_idx - 1 + total_cards) % total_cards
                st.session_state.fc_flipped = False
                trigger_rerun()
        
        with c2:
            # Flip
            label_flip = t('flip').replace('🔄', '').strip()
            if st.button(label_flip, key=f"fc_flip_{current_idx}", icon=":material/sync:", use_container_width=True):
                st.session_state.fc_flipped = not st.session_state.fc_flipped
                trigger_rerun()
        
        with c3:
            # Next
            label_next = t('next_card').replace('➡', '').replace('⬅', '').strip()
            if st.button(label_next, key=f"fc_next_{current_idx}", icon=":material/arrow_forward:", use_container_width=True):
                st.session_state.fc_index = (current_idx + 1) % total_cards
                st.session_state.fc_flipped = False
                trigger_rerun()
        
        with c4:
            # Shuffle
            label_shuffle = t('shuffle').replace('🔀', '').strip()
            if st.button(label_shuffle, key=f"fc_shuffle_{current_idx}", icon=":material/shuffle:", use_container_width=True):
                import random
                import random
                random.shuffle(st.session_state.fc_shuffled_cards)
                st.session_state.fc_index = 0
                st.session_state.fc_flipped = False
                trigger_rerun()

if tab_selection == t("tab_flashcards"):
    render_flashcards()

if tab_selection == t("tab_ranking"):
    sac.divider(label=t("ranking_title"), icon='trophy-fill', align='left', size='lg', color='yellow')
    st.caption(t("ranking_desc"))

    # 自分の学習時間を更新
    if not df_log.empty:
        total_study_minutes = df_log["学習投入時間(分)"].sum()
        total_study_hours = total_study_minutes / 60
        
        # 更新処理
        with st.spinner(t("ranking_updating")):
            st.session_state.sheets_manager.update_ranking(st.session_state.current_user, total_study_hours)
    
    # ランキング取得
    df_rank, err = st.session_state.sheets_manager.get_ranking()
    
    if err:
        st.error(t("ranking_error").format(err))
    else:
        if not df_rank.empty:
            # 自分の順位を確認
            my_rank = df_rank[df_rank["User"] == st.session_state.current_user].index.tolist()
            if my_rank:
                rank_num = my_rank[0] + 1
                st.metric(t("your_rank"), f"{rank_num}{t('rank_suffix')}", f"{df_rank.iloc[my_rank[0]]['TotalStudyTime']:.1f}{t('hours_suffix')}")
            
            # ランキング表示
            st.dataframe(
                df_rank[["User", "TotalStudyTime"]].rename(columns={"User": t("user_label"), "TotalStudyTime": t("study_time_hours")}),
                use_container_width=True,
                hide_index=False
            )
            
            # グラフ表示
            fig_rank = px.bar(df_rank.head(10), x="TotalStudyTime", y="User", orientation='h', 
                              title=t("top_10_users"), text_auto='.1f')
            fig_rank.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_rank, use_container_width=True)
        else:
            st.info(t("no_ranking_data"))

if tab_selection == t("tab_review_notes"):
    sac.divider(label=t('review_notes_title'), icon='journal-bookmark', align='center')
    st.write(t("review_notes_desc"))
    
    if st.session_state.df_notes.empty:
        sac.alert(t("no_notes_yet"), icon='info-circle', color='info')
    else:
        # 検索機能
        st.markdown(f'<div style="margin-bottom:8px; font-weight:600; color:#374151;"><i class="bi bi-search" style="margin-right:6px; color:#3b82f6;"></i>{t("search_problem_id_or_memo")}</div>', unsafe_allow_html=True)
        search_query = st.text_input(t("search"), placeholder=t("enter_keyword"), label_visibility="collapsed")
        
        # フィルタリング
        df_notes_display = st.session_state.df_notes.copy()
        if search_query:
            mask = (df_notes_display["問題ID"].astype(str).str.contains(search_query, case=False, na=False)) | \
                   (df_notes_display["メモ"].astype(str).str.contains(search_query, case=False, na=False))
            df_notes_display = df_notes_display[mask]
        
        st.markdown(f"**{t('total_notes').format(count=len(df_notes_display))}**")
        
        # 表示
        for idx, row in df_notes_display.iterrows():
            with st.expander(f"**{row['問題ID']}** - {row['登録日時']}", expanded=False):
                st.markdown(row['メモ'])
                
                # 削除ボタン
                def delete_note(idx_to_drop):
                    st.session_state.df_notes = st.session_state.df_notes.drop(idx_to_drop).reset_index(drop=True)
                    st.session_state.df_notes.to_csv(user_notes_path, index=False)
                    # sac.alertはrerunしないと消えないため、st.toastを使うか、rerunなしでUI更新を待つ
                    st.toast(t("deleted"), icon="✅")

                st.button(t("delete"), key=f"del_note_{idx}", on_click=delete_note, args=(idx,))

if tab_selection == t("tab_settings"):
    sac.divider(label=t('settings'), icon='gear', align='center')
    st.write(t("settings_desc"))
    
    # 言語設定
    sac.divider(label=t('language_settings_title'), icon='translate', align='left')
    lang = st.selectbox(
        t("display_language"),
        ["日本語", "English", "簡体字"],
        index=["日本語", "English", "簡体字"].index(st.session_state.get("language", "日本語")), key="lang_select"
    )
    if st.session_state.language != lang:
        st.session_state.language = lang
        trigger_rerun()

    # 試験日設定
    sac.divider(label=t('exam_date_settings'), icon='calendar-event', align='left')
    st.caption(t("exam_date_countdown_desc"))
    edate = st.date_input(t("exam_date"), value=st.session_state.exam_date if st.session_state.exam_date else None, key="exam_date_input")
    if st.session_state.exam_date != edate:
        st.session_state.exam_date = edate
        trigger_rerun()

    # ダッシュボード表示設定
    sac.divider(label=t("dashboard_settings"), icon='layout-text-window-reverse', align='left')
    st.caption(t("dashboard_settings_desc"))
    
    widgets_options = ["主要指標", "学習カレンダー", "学習記録", "週間学習プラン", "バッジ"]
    
    widget_name_map = {
        "主要指標": t("widget_metrics"),
        "学習カレンダー": t("widget_calendar"),
        "学習記録": t("widget_log"),
        "週間学習プラン": t("widget_plan"),
        "バッジ": t("widget_badges")
    }
    
    # Ensure defaults are valid
    current_defaults = st.session_state.get("dashboard_widgets_v2", ["主要指標", "学習カレンダー", "学習記録", "週間学習プラン"])
    valid_defaults = [w for w in current_defaults if w in widgets_options]

    selected_widgets = st.multiselect(
        t("select_widgets_label"),
        options=widgets_options,
        default=valid_defaults,
        key="dashboard_widgets_select_v2",
        format_func=lambda x: widget_name_map.get(x, x)
    )
    
    if selected_widgets != st.session_state.get("dashboard_widgets_v2"):
        st.session_state.dashboard_widgets_v2 = selected_widgets
        trigger_rerun()

    # テーマ設定
    sac.divider(label=t('theme_color'), icon='palette', align='left')
    theme_keys = list(THEMES.keys())
    try:
        current_index = theme_keys.index(st.session_state.theme)
    except ValueError:
        current_index = 0
        st.session_state.theme = theme_keys[0]
        
    th = st.selectbox(t("select_theme"), theme_keys, index=current_index, key="theme_select")
    if st.session_state.theme != th:
        st.session_state.theme = th
        trigger_rerun()
    
    # 表示モード設定
    st.markdown("---")
    sac.divider(label=t('display_mode'), icon='moon-stars', align='left')
    st.caption(t("dark_mode_settings_desc"))
    
    display_modes = [t("light_mode"), t("dark_mode"), t("system_setting")]
    current_mode = st.session_state.get("display_mode", t("system_setting"))
    
    try:
        mode_index = display_modes.index(current_mode)
    except ValueError:
        mode_index = 2  # デフォルト: システム設定
    
    selected_mode = st.selectbox(
        t("select_display_mode"),
        display_modes,
        index=mode_index,
        key="display_mode_select"
    )
    
    if st.session_state.get("display_mode") != selected_mode:
        st.session_state.display_mode = selected_mode
        trigger_rerun()

    st.markdown("---")
    keep = st.checkbox(t("keep_input_form_open"), value=st.session_state.get("keep_input_open", True), key="keep_input_open_checkbox")
    st.session_state.keep_input_open = keep
    
    # 週報レポート生成
    sac.divider(label=t('weekly_report_generation'), icon='file-earmark-text', align='left')
    st.caption(t("weekly_report_desc"))
    
    if st.button(t("generate_report"), type="primary", use_container_width=True):
        report = generate_weekly_report(df)
        st.markdown(report, unsafe_allow_html=True)
        
        # ダウンロードボタン
        col_dl1, col_dl2, col_dl3 = st.columns(3)
        
        with col_dl1:
            st.text_area(t("copy_for_clipboard"), value=report, height=200, key="weekly_report_copy")
        
        with col_dl2:
            # PDF出力
            pdf_data = generate_pdf_report(report, st.session_state.current_user, df)
            if pdf_data:
                st.download_button(
                    label=t("download_pdf"),
                    data=pdf_data,
                    file_name=f"weekly_report_{st.session_state.current_user}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            else:
                st.info(t("additional_libs_for_pdf"))
        
        with col_dl3:
            # Excel出力（学習データ）
            excel_data = generate_excel_report(df, st.session_state.current_user)
            if excel_data:
                st.download_button(
                    label=t("download_excel"),
                    data=excel_data,
                    file_name=f"learning_data_{st.session_state.current_user}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.info("Excel出力には追加ライブラリが必要です")
    
    st.markdown("---")
    st.markdown("---")
    st.write(t("future_features"))
    st.write(t("feature_auto_plan"))
    st.write(t("feature_user_tracking"))

st.markdown("</div>", unsafe_allow_html=True)

# ===== カスタム通知（トースト）の表示 =====
if st.session_state.get("show_success_toast", False):
    import time
    toast_id = int(time.time() * 1000)
    st.markdown(f"""
    <style>
    @keyframes slideInFadeOut {{
        0% {{ transform: translateX(100%); opacity: 0; }}
        10% {{ transform: translateX(0); opacity: 1; }}
        80% {{ transform: translateX(0); opacity: 1; }}
        100% {{ transform: translateX(100%); opacity: 0; visibility: hidden; }}
    }}
    .custom-toast-{toast_id} {{
        position: fixed;
        top: 100px;
        right: 20px;
        background-color: #ffffff;
        border-left: 5px solid #10b981;
        padding: 16px 24px;
        border-radius: 8px;
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
        display: flex;
        align-items: center;
        gap: 12px;
        z-index: 10000;
        animation: slideInFadeOut 4s forwards;
    }}
    .toast-icon {{
        color: #10b981;
        font-size: 1.5rem;
    }}
    .toast-message {{
        color: #1f2937;
        font-weight: 600;
        font-size: 1rem;
    }}
    </style>
    <div class="custom-toast-{toast_id}">
        <i class="bi bi-check-circle-fill toast-icon"></i>
        <span class="toast-message">データを追加しました</span>
    </div>
    """, unsafe_allow_html=True)
    # 一度表示したらフラグを下ろす
    st.session_state.show_success_toast = False
